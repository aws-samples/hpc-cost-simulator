#!/usr/bin/env python3
'''
Analyze the jobs parsed from scheduler logs running on an ideal compute cluster

Unlike AnalyzeJobs.py, this only supports parsing the CSV files that are output by the log file parsers.
It produces the same final output as AnalyzeJobs.py.

Copyright Amazon.com, Inc. or its affiliates. All Rights Reserved.
SPDX-License-Identifier: MIT-0
'''

import argparse
from bisect import bisect_left
from config_schema import check_schema
import csv
from ComputeClusterModel.ComputeClusterEvent import ComputeClusterEvent, EndOfHourEvent, InstanceTerminateEvent, JobFinishEvent, ScheduleJobsEvent
from ComputeClusterModel.ComputeInstance import ComputeInstance
from CSVLogParser import CSVLogParser, logger as CSVLogParser_logger
from datetime import datetime, timezone
from JobAnalyzerBase import JobAnalyzerBase, SECONDS_PER_HOUR
from os import path
import json
import logging
from math import ceil, log, log2
from openpyxl import Workbook as XlsWorkbook
from openpyxl.chart import BarChart3D, LineChart as XlLineChart, Reference as XlReference
from openpyxl.styles import Alignment as XlsAlignment, Protection as XlsProtection
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE
from openpyxl.utils import get_column_letter as xl_get_column_letter
from operator import attrgetter, itemgetter
from os import makedirs
from os.path import dirname, realpath
from SchedulerJobInfo import logger as SchedulerJobInfo_logger, SchedulerJobInfo, str_to_timedelta, timestamp_to_datetime
from SchedulerLogParser import logger as SchedulerLogParser_logger
from SortJobs import JobSorter
from sys import exit
from VersionCheck import logger as VersionCheck_logger, VersionCheck

logger = logging.getLogger(__file__)
logger_formatter = logging.Formatter('%(levelname)s:%(asctime)s: %(message)s')
logger_streamHandler = logging.StreamHandler()
logger_streamHandler.setFormatter(logger_formatter)
logger.addHandler(logger_streamHandler)
logger.propagate = False
logger.setLevel(logging.INFO)

debug_level = 0
debug_print_event_summary = False
debug_print_event_details = False
debug_insert_instance = False
debug_remove_instance = False

class ComputeClusterModel(JobAnalyzerBase):

    def __init__(self, csv_parser: CSVLogParser, config_filename: str, output_dir: str, starttime: str, endtime: str, queue_filters: str, project_filters: str) -> None:
        '''
        Constructor

        Args:
            csv_parser (CSVLogParser): Job parser
            config_filename (str): Configuration file
            output_dir (str): Output directory
            starttime (str): Select jobs after the specified time
            endtime (str): Select jobs after the specified time
            queue_filters (str): Queue filters
            project_filters (str): Project filters
        Returns:
            None
        '''
        super().__init__(csv_parser, config_filename, output_dir, starttime, endtime, queue_filters, project_filters)

        if 'ComputeClusterModel' not in self.config:
            self.config['ComputeClusterModel'] = {}
            try:
                validated_config = check_schema(self.config)
            except Exception as e:
                logger.error(f"{config_filename} has errors\n{e}")
                exit(1)
            self.config = validated_config

        if not self.instance_type_info:
            self.get_instance_type_info()

        self._scratch_hourly_stats_filename = path.join(self._output_dir, 'hourly_stats.scratch.csv')
        self._hourly_stats_filename = path.join(self._output_dir, 'hourly_stats.csv')
        self._excel_filename = path.join(self._output_dir, 'hourly_stats.xlsx')

        self._boot_time_td = str_to_timedelta(self.config['ComputeClusterModel']['BootTime'])
        self._idle_time_td = str_to_timedelta(self.config['ComputeClusterModel']['IdleTime'])

        self._scheduling_frequency_td = str_to_timedelta(self.config['ComputeClusterModel']['SchedulingFrequency'])
        self._share_instances = self.config['ComputeClusterModel']['ShareInstances']
        self._max_instance_types = False

        self._instance_families = []
        for instance_type in self.instance_types:
            instance_family = instance_type.split(r'.')[0]
            if instance_family not in self._instance_families:
                self._instance_families.append(instance_family)
        self._instance_families.sort()
        logger.debug(f"instance_families: {self._instance_families}")

        logger.debug(f"{len(self._instance_families)} instance families and {len(self.instance_types)} instance types")

        if self._max_instance_types:
            self.instance_types = []
            for instance_family in self._instance_families:
                self.instance_types.append(self.instance_family_info[instance_family]['MaxInstanceType'])
            self.instance_types.sort()

        # Sort the instance types in instance_family_info by CoreCount
        for instance_family in self.instance_family_info:
            self.instance_family_info[instance_family]['instance_types'].sort(key=lambda instance_family: self.instance_type_info[instance_family]['DefaultCores'])

        self._find_best_instance_families()

        self._instance_families_used = []

        self.current_time = timestamp_to_datetime(0)

        self._unscheduled_jobs = {}
        for pricing_option in ComputeInstance.PRICING_OPTIONS:
            self._unscheduled_jobs[pricing_option] = []
        self._schedule_jobs_event = None

        self._events = []

        self.total_jobs = 0
        self.total_failed_jobs = 0

        self._hourly_event = None
        self._number_of_hours = 0

        # Create the data structure for saving instances
        self._create_instance_lists()

        # Need to keep a list of the termination events in case we need to cancel them when a job gets scheduled on an idle instance
        # Index is the instance id.
        self._instance_termination_events = {}

        self._init_scratch_csv()
        self._init_hourly_costs()

    def schedule_jobs(self):
        '''
        Model how jobs get scheduled on ideal Compute Cluster.

        Jobs must first be sorted by eligible_time.

        Analyze jobs 1 by 1.
        Select a pricing plan and an instance type for each job and calculate the job's cost.
        Bin jobs by the hour in which they started and save the hourly jobs in separate CSV files.
        Process the hourly files to accumulate the hourly on-demand and spot costs broken out by instance family.
        Write the hourly costs into a separate CSV file.
        Write an overall job summary.

        Scalability is a key consideration because millions of jobs must be processed so the analysis cannot be
        done in memory.
        First breaking the jobs out into hourly chunks makes the process scalable.
        '''
        logger.info(f"Scheduling jobs on the compute cluster.")
        self._clear_job_stats()   # allows calling multiple times in testing

        previous_eligible_time = None
        while True:
            job = self._scheduler_parser.parse_job()
            if not job:
                logger.info(f"No more jobs to schedule")
                break

            # Check if ran between starttime and endtime
            assert self._scheduler_parser._job_in_time_window(job) # nosec

            self.total_jobs += 1
            if previous_eligible_time and (job.eligible_time_dt < previous_eligible_time):
                logger.error(f"Jobs are not sorted by eligible_time. Job {job.job_id} eligible_time={job.eligible_time_dt} < {previous_eligible_time}")
                exit(2)
            previous_eligible_time = job.eligible_time_dt
            if self.current_time < job.eligible_time_dt:
                logger.debug(f"{self.current_time}: Advance time to {job.eligible_time_dt} to schedule job {job.job_id}")
                self.advance_time(job.eligible_time_dt)
            job_runtime_minutes = job.run_time_td.total_seconds()/60
            spot_threshold = self.config['consumption_model_mapping']['maximum_minutes_for_spot']
            if job_runtime_minutes < spot_threshold:
                pricing_option = ComputeInstance.SPOT
            else:
                pricing_option = ComputeInstance.ON_DEMAND
            self._add_job_to_collector(job)
            self.submit_job(job, pricing_option)
        self.finish_jobs()
        self.write_stats()
        logger.info(f"Finished processing {self.total_jobs-self.total_failed_jobs}/{self.total_jobs} jobs")

    def advance_time(self, new_time: datetime) -> None:
        while self._events and self._events[0].time < new_time:
            self._process_event()
        assert new_time >= self.current_time, f"{new_time} < {self.current_time}" # nosec
        self.current_time = new_time

    def submit_job(self, job: SchedulerJobInfo, pricing_option: str) -> None:
        logger.debug(f"{self.current_time}: Submit job {job.job_id} cores={job.num_cores} max_mem={job.max_mem_gb}GB")

        # If SchedulingFrequency == 0 then schedule immediately
        if self._scheduling_frequency_td.total_seconds() == 0:
            self._schedule_job(job, pricing_option, allow_new_instance=True)
            return

        # First schedule the job immediately if there is a free instance
        if self._schedule_job(job, pricing_option, allow_new_instance=False):
            return
        # This throttles job scheduling to allow instances to be grouped to run more than 1 job.
        self._unscheduled_jobs[pricing_option].append(job)
        if not self._schedule_jobs_event:
            self._schedule_jobs_event = ScheduleJobsEvent(self.current_time + self._scheduling_frequency_td)
            self._insert_event(self._schedule_jobs_event)

    def finish_jobs(self) -> None:
        '''
        Finish executing all remaining events
        '''
        while self._events:
            self._process_event()

    def write_stats(self) -> None:
        logger.info(f"Writing out final stats")
        self._init_csv()
        self._init_excel()
        self._process_scratch_csv()
        self._finish_excel()

    def _find_best_instance_families(self):
        '''
        Find the best instance families for each memory to core ratio.

        The memory/core ratio on EC2 instances is always a power of 2.
        So, if you bin each job according to its closest EC2 mem/core ration then instance family selection is easy for a job.
        The cost/core for each instance type is always the same, so we can always choose the cheapest instance family based soley on mem/core of the job.

        When in
        '''
        logger.debug("_find_best_instance_families")
        self._best_instance_family = {}
        for pricing_option in ComputeInstance.PRICING_OPTIONS:
            logger.debug(f"    {pricing_option}")
            self._best_instance_family[pricing_option] = {}
            for instance_family in self._instance_families:
                instance_type = self.instance_family_info[instance_family]['MaxInstanceType']
                number_of_cores = self.instance_type_info[instance_type]['DefaultCores']
                memory_in_gb = self.instance_type_info[instance_type]['MemoryInMiB'] / 1024
                memory_per_core = int(round(memory_in_gb / number_of_cores, 0))
                cost = self._get_instance_type_pricing(instance_type, pricing_option)
                if cost == None:
                    continue
                cost_per_core = cost / number_of_cores
                self.instance_family_info[instance_family]['cost_per_core'] = {pricing_option: cost_per_core}
                if memory_per_core not in self._best_instance_family[pricing_option]:
                    logger.debug(f"        {memory_per_core}: {instance_family}: {cost_per_core}")
                    self._best_instance_family[pricing_option][memory_per_core] = instance_family
                else:
                    best_instance_family_so_far = self._best_instance_family[pricing_option][memory_per_core]
                    best_cost_per_core_so_far = self.instance_family_info[best_instance_family_so_far]['cost_per_core'][pricing_option]
                    if cost_per_core < best_cost_per_core_so_far:
                        logger.debug(f"        {memory_per_core}: {instance_family}: {cost_per_core}")
                        self._best_instance_family[pricing_option][memory_per_core] = instance_family
            self._bucket_list = sorted(self._best_instance_family[pricing_option].keys())

    def _get_instance_type_pricing(self, instance_type: str, pricing_option: str) -> float:
        assert instance_type in self.instance_types, f"{instance_type} not in {self.instance_types}" # nosec
        assert pricing_option in ComputeInstance.PRICING_OPTIONS, f"{pricing_option} not in {ComputeInstance.PRICING_OPTIONS}" # nosec
        if pricing_option == ComputeInstance.SPOT:
            # hpc instances don't have spot pricing
            return self.instance_type_info[instance_type]['pricing'][pricing_option].get('max', None)
        else:
            return self.instance_type_info[instance_type]['pricing'][pricing_option]

    def _schedule_job(self, job: SchedulerJobInfo, pricing_option: str, allow_new_instance: bool) -> bool:
        '''
        Schedule job to run on the cluster

        Allocate job to existing instance or to a new instance.
        Create JobFinish event for the job.
        If new instance, then JobFinish event should include boot time.

        Need to figure out the optimal instance selection to reduce costs.

        Max sized instances:
        One possibility is to use the maximum instance size in an instance family.
        This may allow better job packing, reduce boot times, reduce idle instances.
        The down side is that the utilization of the instance is likely to be lower.

        Pick least expensive instance:
        This seems obvious. Pick the cheapest instance that meets job requirements.

        Pick the instance with highest utilization:
        The thought here is to maximize the utilization of each instance and allow instances
        with lower utilization to go idle and get terminated.

        Pick the instance with lowest utilization:
        This might improve job performance, but would tend to keep lightly utilized instances
        to keep running.
        This is likely to be more expensive.
        '''
        assert pricing_option in ComputeInstance.PRICING_OPTIONS # nosec
        compute_instance = self._allocate_instance(job, pricing_option, allow_new_instance)
        if not compute_instance:
            return False
        if compute_instance.running_time > job.start_time_dt:
            job.start_time_dt = compute_instance.running_time
        else:
            job.start_time_dt = self.current_time
        job.wait_time_td = job.start_time_dt - self.current_time
        job.finish_time_dt = job.start_time_dt + job.run_time_td
        job_finish_event = JobFinishEvent(job.finish_time_dt, job, compute_instance)
        logger.debug(f"{self.current_time}: Scheduled job {job.job_id} start={job.start_time_dt} finish={job.finish_time_dt} elapsed={job.run_time_td}")
        self._insert_event(job_finish_event)
        if not self._hourly_event:
            self._add_hourly_event()
        return True

    def _schedule_jobs(self) -> None:
        '''
        Scheduler loop

        Schedule all unscheduled jobs.
        Reverse sort unscheduled jobs by memory/core, number of cores and schedule them.

        First try to schedule them on instances that have available resources.
        Then start new instances to jobs that don't fit on existing instances.
        The point of doing this is to try to allow multiple jobs to get scheduled on 1 instance so that they are used more efficiently.

        One possible optimization is to combine instances of the same size into a larger instance.
        But, I'm skeptical that this will reduce costs or save time because the unused portion of the instance will be unused, but the instance
        won't be idle until all jobs complete.
        '''
        logger.debug(f"{self.current_time}: _schedule_jobs")
        self._schedule_jobs_event = None
        num_unscheduled_jobs = 0
        for pricing_option in ComputeInstance.PRICING_OPTIONS:
            num_unscheduled_jobs += len(self._unscheduled_jobs[pricing_option])
        logger.debug(f"    {num_unscheduled_jobs} unscheduled jobs")
        if num_unscheduled_jobs == 0:
            return
        self._schedule_jobs_event = ScheduleJobsEvent(self.current_time + self._scheduling_frequency_td)
        self._insert_event(self._schedule_jobs_event)

        for pricing_option in ComputeInstance.PRICING_OPTIONS:
            if self._unscheduled_jobs[pricing_option]:
                logger.debug(f"    {pricing_option}")
            jobs_to_schedule = self._unscheduled_jobs[pricing_option]
            self._unscheduled_jobs[pricing_option] = []
            # jobs_to_schedule.sort(key=lambda job: (-job.core_count))
            for job in jobs_to_schedule:
                if self._schedule_job(job, pricing_option, allow_new_instance=True):
                    logger.debug(f"            Scheduled {job} on existing instance")
                else:
                    logger.error(f"Could not schedule job {job.job_id}")

    def _get_memory_bucket(self, job: SchedulerJobInfo) -> int:
        mem_per_core = int(ceil(job.max_mem_gb / job.num_cores))
        pos = bisect_left(self._bucket_list, mem_per_core)
        assert pos < len(self._bucket_list) # nosec
        return self._bucket_list[pos]

    def _create_instance_lists(self):
        '''
        Group instances by pricing_option, cost per core, and utilization
        '''
        self._number_of_instances = {}
        self._instances = {}
        self._instances_sorted = {}
        for pricing_option in ComputeInstance.PRICING_OPTIONS:
            self._number_of_instances[pricing_option] = 0
            self._instances[pricing_option] = []
            self._instances_sorted[pricing_option] = True

    def _get_instance_count(self, pricing_options=ComputeInstance.PRICING_OPTIONS):
        total_instances = 0
        for pricing_option in pricing_options:
            total_instances += self._number_of_instances[pricing_option]
        return total_instances

    def _get_instance_list(self, pricing_option: str) -> [ComputeInstance]:
        assert pricing_option in ComputeInstance.PRICING_OPTIONS # nosec
        return self._instances[pricing_option]

    def _add_instance(self, instance: ComputeInstance) -> None:
        self._number_of_instances[instance.pricing_option] += 1
        self._insert_instance(instance)

    def _delete_instance(self, instance: ComputeInstance) -> None:
        self._number_of_instances[instance.pricing_option] -= 1
        self._remove_instance(instance)

    def _insert_instance(self, instance: ComputeInstance) -> None:
        self._instances[instance.pricing_option].append(instance)
        self._instances_sorted[instance.pricing_option] = False
        if debug_insert_instance:
            self._print_instances(instance.pricing_option)

    def _sort_instances(self, pricing_option: str) -> None:
        if self._instances_sorted[pricing_option]:
            return
        # self._instances[pricing_option].sort(key=attrgetter('cost_per_core'))
        # self._instances[pricing_option].sort(key=attrgetter('cost_per_core', 'utilization'))
        self._instances[pricing_option].sort(key=attrgetter('cost_per_core', 'unutilization'))
        # self._instances[pricing_option].sort(key=attrgetter('unutilization'))
        # self._instances[pricing_option].sort(key=attrgetter('unutilization', 'cost_per_core'))
        # self._instances[pricing_option].sort(key=attrgetter('utilization'))
        #self._instances[pricing_option].sort(key=attrgetter('utilization', 'cost_per_core'))
        self._instances_sorted[pricing_option] = True

    def _remove_instance(self, instance: ComputeInstance) -> None:
        self._instances[instance.pricing_option].remove(instance)
        if debug_remove_instance:
            self._print_instances(instance.pricing_option)

    def _change_utilization(self, instance: ComputeInstance) -> None:
        '''
        Move the instance in the list after its utilization is updated
        '''
        self._remove_instance(instance)
        self._insert_instance(instance)

    def _print_instances(self, pricing_option='', summary_only=False):
        if pricing_option == '':
            pricing_options = ComputeInstance.PRICING_OPTIONS
        else:
            pricing_options = [pricing_option]
        instance_count = {}
        instance_count_by_instance_type = {}
        for pricing_option in pricing_options:
            instance_count[pricing_option] = {'total': 0, 'idle': 0}
            instance_count_by_instance_type[pricing_option] = {}
            for instance in self._instances[pricing_option]:
                idle = len(instance.jobs) == 0
                instance_count[pricing_option]['total'] += 1
                if instance.instance_type not in instance_count_by_instance_type[pricing_option]:
                    instance_count_by_instance_type[pricing_option][instance.instance_type] = {'total': 0, 'idle': 0, 'instances': []}
                instance_count_by_instance_type[pricing_option][instance.instance_type]['total'] += 1
                if idle:
                    instance_count[pricing_option]['idle'] += 1
                    instance_count_by_instance_type[pricing_option][instance.instance_type]['idle'] += 1
                instance_count_by_instance_type[pricing_option][instance.instance_type]['instances'].append(instance)
        if not summary_only:
            logger.debug(f"{self.current_time}: Instance details:")
            for pricing_option in pricing_options:
                logger.debug(f"    {pricing_option}: total={instance_count[pricing_option]['total']} idle={instance_count[pricing_option]['idle']}")
                for instance_type in sorted(instance_count_by_instance_type[pricing_option].keys()):
                    logger.debug(f"        {instance_type}: total={instance_count_by_instance_type[pricing_option][instance_type]['total']} idle={instance_count_by_instance_type[pricing_option][instance_type]['idle']}")
                    for instance in instance_count_by_instance_type[pricing_option][instance_type]['instances']:
                        logger.debug(f"            {instance}")
                        for job in instance.jobs:
                            logger.debug(f"                job {job.job_id}: start={job.start_time_dt} finish={job.finish_time_dt}")
        logger.debug(f"{self.current_time}: Instance summary:")
        for pricing_option in pricing_options:
            logger.debug(f"    {pricing_option}: total={instance_count[pricing_option]['total']} idle={instance_count[pricing_option]['idle']}")
            for instance_type in sorted(instance_count_by_instance_type[pricing_option].keys()):
                logger.debug(f"        {instance_type}: total={instance_count_by_instance_type[pricing_option][instance_type]['total']} idle={instance_count_by_instance_type[pricing_option][instance_type]['idle']}")

    def _allocate_instance(self, job: SchedulerJobInfo, pricing_option: str, allow_new_instance: bool) -> ComputeInstance:
        assert pricing_option in ComputeInstance.PRICING_OPTIONS # nosec
        compute_instance = None
        if self._share_instances:
            compute_instance = self._allocate_existing_instance(job, pricing_option)
            if compute_instance or not allow_new_instance:
                return compute_instance
        compute_instance = self._allocate_new_instance(job, pricing_option)
        if not compute_instance and pricing_option == ComputeInstance.SPOT:
            compute_instance = self._allocate_new_instance(job, ComputeInstance.ON_DEMAND)
        return compute_instance

    def _allocate_existing_instance(self, job: SchedulerJobInfo, pricing_option: str) -> ComputeInstance:
        assert pricing_option in ComputeInstance.PRICING_OPTIONS # nosec
        if not self._share_instances:
            return None
        self._sort_instances(pricing_option)
        for instance in self._get_instance_list(pricing_option):
            current_utilization = instance.utilization
            if current_utilization == 1.0:
                continue
            try:
                instance.allocate_job(job, pricing_option, self.current_time)
            except AssertionError:
                continue
            self._change_utilization(instance)
            logger.debug(f"{self.current_time}: Allocated existing {instance} for job {job.job_id}")
            if instance.instance_id in self._instance_termination_events:
                self._delete_termination_event(instance)
            return instance
        return None

    def _allocate_new_instance(self, job: SchedulerJobInfo, pricing_option: str) -> ComputeInstance:
        assert pricing_option in ComputeInstance.PRICING_OPTIONS # nosec
        cheapest_instance_type = None
        lowest_hourly_cost = None
        for instance_type in self.instance_types:
            if (self.instance_type_info[instance_type]['MemoryInMiB']/1024) < job.max_mem_gb:
                continue
            if (self.instance_type_info[instance_type]['DefaultCores']) < job.num_cores:
                continue
            if pricing_option not in self.instance_type_info[instance_type]['pricing']:
                continue
            hourly_cost = self._get_instance_type_pricing(instance_type, pricing_option)
            if hourly_cost == None:
                continue
            if not cheapest_instance_type or (hourly_cost < lowest_hourly_cost):
                cheapest_instance_type = instance_type
                lowest_hourly_cost = hourly_cost
        if not cheapest_instance_type:
            if pricing_option == ComputeInstance.SPOT:
                logger.warning(f"Couldn't find {pricing_option} instance type with {job.num_cores} cores and {job.max_mem_gb} GB memory to run job {job.job_id}")
            else:
                logger.error(f"Couldn't find {pricing_option} instance type with {job.num_cores} cores and {job.max_mem_gb} GB memory to run job {job.job_id}")
            return None
        instance_type = cheapest_instance_type
        compute_instance = self._start_instance(cheapest_instance_type, pricing_option)
        compute_instance.allocate_job(job, pricing_option, self.current_time)
        self._change_utilization(compute_instance)
        if compute_instance.instance_id in self._instance_termination_events:
            self._delete_termination_event(instacompute_instancence)
        return compute_instance

    def _start_instance(self, instance_type: str, pricing_option: str) -> ComputeInstance:
        assert instance_type in self.instance_types, f"{instance_type} not in {self.instance_types}" # nosec
        assert pricing_option in ComputeInstance.PRICING_OPTIONS, f"{pricing_option} not in {ComputeInstance.PRICING_OPTIONS}" # nosec
        instance_type_info = self.instance_type_info[instance_type]
        compute_instance = ComputeInstance('EC2', instance_type, instance_type_info['DefaultCores'], instance_type_info['DefaultThreadsPerCore'], ht_enabled=False, mem_gb=instance_type_info['MemoryInMiB']/1024, pricing_option=pricing_option, hourly_cost=self._get_instance_type_pricing(instance_type, pricing_option), start_time=self.current_time, running_time=self.current_time + self._boot_time_td)
        logger.debug(f"{self.current_time}: Started {compute_instance}")
        self._add_instance(compute_instance)
        return compute_instance

    def _insert_event(self, event: ComputeClusterEvent):
        assert event.time >= self.current_time, f"{event.time} < {self.current_time} {event}" # nosec

        self._events.append(event)
        self._events.sort(key=attrgetter('time'))
        self._print_events()
        return

    def _print_events(self):
        if not debug_print_event_summary:
            return
        logger.debug(f"{self.current_time}: {len(self._events)} Events:")
        if not debug_print_event_details:
            return
        previous_event_time = self.current_time
        for idx, event in enumerate(self._events):
            logger.debug(f"    {idx}, {event.time}: {event}")
            assert event.time >= previous_event_time, f"{event.time} < {previous_event_time}" # nosec
            previous_event_time = event.time

    def _process_event(self):
        event = self._events.pop(0)
        assert event.time >= self.current_time, f"{event.time} < {self.current_time}" # nosec
        self.current_time = event.time
        if isinstance(event, ScheduleJobsEvent):
            self._schedule_jobs()
        elif isinstance(event, EndOfHourEvent):
            logger.debug(f"{self.current_time}: Processing EndOfHourEvent")
            self._hourly_event = None

            total_instances = self._get_instance_count()
            for pricing_option in ComputeInstance.PRICING_OPTIONS:
                for instance in self._get_instance_list(pricing_option):
                    self._update_hourly_costs(instance)
            self._write_scratch_csv_row()
            self._init_hourly_costs()

            if total_instances and len(self._events) == 0:
                self._print_instances()
                raise RuntimeError(f"{total_instances} instances running without any events")

            # If there are any future events then set another event for the end of the hour
            # Record the hourly costs even if the cluster is idle. This is because if savings plans are used then there will be costs.
            if total_instances or len(self._events) > 0:
                self._add_hourly_event()
                # Print out heart beat message every 24 hours
                if (self._relative_hour % 24) == 0:
                    logger.info(f"Finished processing jobs for hour {self._current_hour} = {timestamp_to_datetime(self._current_hour * SECONDS_PER_HOUR)}")
                    logger.info(f"    {self.total_jobs} jobs scheduled")
            else:
                logger.info("Finished processing events")
                logger.info(f"    {self.total_jobs} jobs scheduled")
        elif isinstance(event, JobFinishEvent):
            logger.debug(f"{event.time}: Job {event.job.job_id} finished")
            instance = event.instance
            current_utilization = instance.utilization
            instance.finish_job(event.job)
            self._change_utilization(instance)
            if not instance.jobs:
                self._add_termination_event(instance, self.current_time + self._idle_time_td)
        elif isinstance(event, InstanceTerminateEvent):
            instance = event.instance
            del self._instance_termination_events[instance.instance_id]
            if instance.jobs:
                logger.debug(f"{self.current_time}: Skipping terminate {instance.instance_type} {instance.pricing_option} because running {len(instance.jobs)} jobs running")
            else:
                logger.debug(f"{self.current_time}: Terminate {instance}")
                try:
                    self._delete_instance(instance)
                except ValueError:
                    logger.error(f"Couldn't remove {instance}")
                    self._print_instances(instance.pricing_option)
                    raise
                self._update_hourly_costs(instance)
        else:
            raise RuntimeError(f"Unsupported event: {event}")

    def _add_hourly_event(self):
        logger.debug(f"{self.current_time}: Adding EndOfHourEvent")
        self._current_hour = int(self.current_time.timestamp() // SECONDS_PER_HOUR)
        self._current_hour_dt = timestamp_to_datetime(self._current_hour * SECONDS_PER_HOUR)
        logger.debug(f"    _current_hour={self._current_hour}={self._current_hour_dt} dst={self._current_hour_dt.dst()}")
        assert self._current_hour_dt <= self.current_time, f"{self._current_hour_dt} > {self.current_time}" # nosec
        if not self._first_hour:
            self._first_hour = self._current_hour
            logger.info(f"First hour: {self._first_hour} = {timestamp_to_datetime(self._first_hour * SECONDS_PER_HOUR)}")
            self._init_hourly_costs()
        self._relative_hour = self._current_hour - self._first_hour
        self._next_hour = self._current_hour + 1
        self._next_hour_dt = timestamp_to_datetime(self._next_hour * SECONDS_PER_HOUR)
        logger.debug(f"    _relative_hour={self._relative_hour}")
        logger.debug(f"    _next_hour={self._next_hour}={self._next_hour_dt} dst={self._next_hour_dt.dst()}")
        self._hourly_event = EndOfHourEvent(self._next_hour_dt)
        self._insert_event(self._hourly_event)

    def _add_termination_event(self, instance: ComputeInstance, termination_time: datetime) -> None:
        assert instance.instance_id not in self._instance_termination_events, f"{instance.instance_id} in {self._instance_termination_events}" # nosec
        event = InstanceTerminateEvent(termination_time, instance)
        self._insert_event(event)
        self._instance_termination_events[instance.instance_id] = event

    def _delete_termination_event(self, instance: ComputeInstance) -> None:
        assert instance.instance_id in self._instance_termination_events, f"{instance.instance_id} not in {self._instance_termination_events}" # nosec
        termination_event = self._instance_termination_events[instance.instance_id]
        logger.debug(f"{self.current_time}: Cancelling {termination_event}")
        self._events.remove(termination_event)
        del self._instance_termination_events[instance.instance_id]

    def _init_scratch_csv(self):
        '''
        Create a scratch CSV file with data for all configured instance families.

        This is done for scalability reasons.
        We don't know which instance families will actually be used so write a scratch CSV file with data with all CSV files
        so we don't have to store the data in memory.

        When the jobs are complete go back and read the scratch file and write out a file with only the instance families that were used
        to reduce the total number of columns.
        '''
        self._scratch_hourly_stats_fh = open(self._scratch_hourly_stats_filename, 'w')
        self._scratch_hourly_stats_csv_writer = csv.writer(self._scratch_hourly_stats_fh, dialect='excel')
        field_names = ['Relative Hour', 'OnDemand Core Hours', 'Total OnDemand Costs', 'Spot Core Hours', 'Spot Costs']
        for instance_family in self._instance_families:
            field_names.append(f"{instance_family} Core Hours")
            field_names.append(f"{instance_family} OnDemand Costs")
        self._scratch_hourly_stats_csv_writer.writerow(field_names)
        self._scratch_hourly_stats_fh.flush()

    def _init_csv(self):
        self._scratch_hourly_stats_fh.close()
        self._scratch_hourly_stats_fh = open(self._scratch_hourly_stats_filename, 'r')
        self._scratch_hourly_stats_csv_reader = csv.reader(self._scratch_hourly_stats_fh, dialect='excel')
        next(self._scratch_hourly_stats_csv_reader)

        self._hourly_stats_fh = open(self._hourly_stats_filename, 'w')
        self._hourly_stats_csv_writer = csv.writer(self._hourly_stats_fh, dialect='excel')
        field_names = ['Relative Hour', 'OnDemand Core Hours', 'Total OnDemand Costs', 'Spot Core Hours', 'Spot Costs']
        self._instance_families_used.sort()
        for instance_family in self._instance_families_used:
            field_names.append(f"{instance_family} Core Hours")
            field_names.append(f"{instance_family} OnDemand Costs")
        self._hourly_stats_csv_writer.writerow(field_names)

    def _init_hourly_costs(self) -> None:
        self._hourly_costs = {}
        for pricing_option in ComputeInstance.PRICING_OPTIONS:
            self._hourly_costs[pricing_option] = {
                'core hours': 0.0,
                'total costs': 0.0,
            }
        for instance_family in self._instance_families:
            self._hourly_costs[ComputeInstance.ON_DEMAND][instance_family] = {
                'core hours': 0.0,
                'costs': 0.0
            }

    def _update_hourly_costs(self, instance: ComputeInstance) -> None:
        assert self._current_hour_dt <= self.current_time, f"{self._current_hour_dt} > {self.current_time}" # nosec
        hours_td = self.current_time - max(instance.start_time, self._current_hour_dt)
        hours = hours_td.total_seconds() / SECONDS_PER_HOUR
        assert 0.0 <= hours <= 1.0, f"invalid instance hours: {hours} > 1.0 or < 0.0\n    current time: {self.current_time}\n    current hour: {self._current_hour_dt}\n    {instance}" # nosec
        core_hours = hours * instance.number_of_cores
        cost = hours * instance.hourly_cost
        self._hourly_costs[instance.pricing_option]['core hours'] += core_hours
        self._hourly_costs[instance.pricing_option]['total costs'] += cost
        if instance.pricing_option == ComputeInstance.ON_DEMAND:
            instance_family = instance.instance_type.split('.')[0]
            if instance_family not in self._instance_families_used:
                self._instance_families_used.append(instance_family)
            self._hourly_costs[instance.pricing_option][instance_family]['core hours'] += core_hours
            self._hourly_costs[instance.pricing_option][instance_family]['costs'] += cost

    def _write_scratch_csv_row(self) -> None:
        if self._relative_hour < 0:
            # Discard information for hours before starttime
            logger.debug(f"Discarding hourly stats for relative hour {self._relative_hour}")
            return
        if self._last_hour and self._current_hour > self._last_hour:
            # Discard information for hours after endtime
            logger.debug(f"Discarding hourly stats for relative hour {self._relative_hour}")
            return
        field_values = [self._relative_hour]
        for pricing_option in ComputeInstance.PRICING_OPTIONS:
            field_values.append(self._hourly_costs[pricing_option]['core hours'])
            field_values.append(self._hourly_costs[pricing_option]['total costs'])
        for instance_family in self._instance_families:
            field_values.append(self._hourly_costs[ComputeInstance.ON_DEMAND][instance_family]['core hours'])
            field_values.append(self._hourly_costs[ComputeInstance.ON_DEMAND][instance_family]['costs'])
        self._scratch_hourly_stats_csv_writer.writerow(field_values)
        self._scratch_hourly_stats_fh.flush()
        self._number_of_hours += 1

    def _read_scratch_csv(self) -> bool:
        try:
            scratch_csv_field_values = next(self._scratch_hourly_stats_csv_reader)
        except StopIteration:
            return False
        self._relative_hour = int(scratch_csv_field_values.pop(0))
        self._current_hour = self._relative_hour + self._first_hour
        self._next_hour = self._current_hour + 1
        self._hourly_costs = {}
        for pricing_option in ComputeInstance.PRICING_OPTIONS:
            self._hourly_costs[pricing_option] = {}
            self._hourly_costs[pricing_option]['core hours'] = float(scratch_csv_field_values.pop(0))
            self._hourly_costs[pricing_option]['total costs'] = float(scratch_csv_field_values.pop(0))
        for instance_family in self._instance_families:
            self._hourly_costs[ComputeInstance.ON_DEMAND][instance_family] = {}
            self._hourly_costs[ComputeInstance.ON_DEMAND][instance_family]['core hours'] = float(scratch_csv_field_values.pop(0))
            self._hourly_costs[ComputeInstance.ON_DEMAND][instance_family]['costs'] = float(scratch_csv_field_values.pop(0))
        return True

    def _write_csv_row(self) -> None:
        field_values = [self._relative_hour]
        for pricing_option in ComputeInstance.PRICING_OPTIONS:
            field_values.append(self._hourly_costs[pricing_option]['core hours'])
            field_values.append(self._hourly_costs[pricing_option]['total costs'])
        for instance_family in self._instance_families_used:
            field_values.append(self._hourly_costs[ComputeInstance.ON_DEMAND][instance_family]['core hours'])
            field_values.append(self._hourly_costs[ComputeInstance.ON_DEMAND][instance_family]['costs'])
        self._hourly_stats_csv_writer.writerow(field_values)

    def _init_excel(self) -> None:
        '''
        Create Excel file
        '''
        logger.info('')
        logger.info(f"Writing hourly stats to {self._excel_filename}")

        self._excel_wb = XlsWorkbook()
        xls_locked = XlsProtection(locked=True)
        xls_unlocked = XlsProtection(locked=False)

        # Create worksheets
        excel_summary_ws = self._excel_wb.active
        excel_summary_ws.title = 'CostSummary'
        #excel_summary_ws.protection.sheet = xls_locked

        excel_instance_family_summary_ws = self._excel_wb.create_sheet(title='InstanceFamilySummary')
        excel_instance_family_summary_ws.protection.sheet = xls_locked

        excel_job_stats_ws = self._excel_wb.create_sheet(title='JobStats')
        excel_job_stats_ws.protection.sheet = xls_locked

        excel_config_ws = self._excel_wb.create_sheet(title='Config')
        excel_config_ws.protection.sheet = xls_locked

        excel_instance_info_ws = self._excel_wb.create_sheet(title='InstanceFamilyRates')
        excel_instance_info_ws.protection.sheet = xls_locked

        self._excel_hourly_ws = self._excel_wb.create_sheet(title='Hourly')
        self._excel_hourly_ws.protection.sheet = xls_locked
        excel_core_hours_chart_ws = self._excel_wb.create_sheet(title='Core Hours Chart')

        # CostSummary Worksheet
        excel_summary_ws.column_dimensions['A'].width = 35
        excel_summary_ws.column_dimensions['B'].width = 25
        excel_summary_ws.column_dimensions['B'].alignment = XlsAlignment(horizontal='right')
        row = 1
        excel_summary_ws[f'A{row}'] = 'First hour to analyze'
        first_hour_cell = excel_summary_ws[f'B{row}']
        first_hour_cell.value = 0
        first_hour_cell.protection = xls_unlocked
        first_hour_cell_ref = f'CostSummary!$B${row}'
        row += 1
        excel_summary_ws[f'A{row}'] = 'Last hour to analyze'
        last_hour_cell = excel_summary_ws[f'B{row}']
        last_hour_cell.protection = xls_unlocked
        last_hour_cell.value = self._number_of_hours - 1
        last_hour_cell_ref = f'CostSummary!$B${row}'
        row += 2
        excel_summary_ws.cell(row=row, column=1).value = f'EC2 Savings Plan (ESP) Hourly Commits:'
        esp_hourly_commit_cell_refs = {}
        esp_hourly_commit_first_row = row + 1
        for instance_family in self._instance_families_used:
            row += 1
            excel_summary_ws[f'A{row}'] = f'{instance_family}'
            cell = excel_summary_ws[f'B{row}']
            cell.value = 0
            cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
            cell.protection = xls_unlocked
            esp_hourly_commit_cell_refs[instance_family] = f'CostSummary!$B${row}'
        esp_hourly_commit_last_row = row
        row += 1
        excel_summary_ws[f'A{row}'] = 'Total'
        cell = excel_summary_ws[f'B{row}']
        if self._instance_families_used:
            cell.value = f"=sum(B{esp_hourly_commit_first_row}:B{esp_hourly_commit_last_row})"
        else:
            cell.value = 0
        cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
        self._esp_hourly_commit_cell_ref = f"CostSummary!${cell.column_letter}${cell.row}"
        row += 2
        excel_summary_ws[f'A{row}'] = 'CSP Hourly Commit'
        cell = excel_summary_ws[f'B{row}']
        cell.value = 0
        cell.protection = xls_unlocked
        cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
        self._csp_hourly_commit_cell_ref = f'CostSummary!${cell.column_letter}${cell.row}'
        row += 2
        excel_summary_ws[f'A{row}'] = 'Total Spot'
        self._total_spot_cell = excel_summary_ws[f'B{row}']
        self._total_spot_cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
        row += 1
        excel_summary_ws[f'A{row}'] = 'Total OD'
        self._total_od_cell = excel_summary_ws[f'B{row}']
        self._total_od_cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
        row += 1
        excel_summary_ws[f'A{row}'] = 'Total ESP'
        self._total_esp_cell = excel_summary_ws[f'B{row}']
        self._total_esp_cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
        row += 1
        excel_summary_ws[f'A{row}'] = 'Total CSP'
        self._total_csp_cell = excel_summary_ws[f'B{row}']
        self._total_csp_cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
        row += 1
        excel_summary_ws[f'A{row}'] = 'Total'
        total_cell = excel_summary_ws[f'B{row}']
        total_cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
        total_cell_ref = f'CostSummary!${total_cell.column_letter}${total_cell.row}'
        row += 2
        excel_summary_ws.cell(row=row, column=1).value = 'Use Excel Solver to optimize savings plans'
        row += 1
        excel_summary_ws.cell(row=row, column=1).value = 'Enable solver'
        row += 1
        excel_summary_ws.cell(row=row, column=1).value = 'File -> Options'
        row += 1
        excel_summary_ws.cell(row=row, column=1).value = 'Select Add-ins on the left'
        row += 1
        excel_summary_ws.cell(row=row, column=1).value = 'Manage: Excel Add-ins, Click Go'
        row += 1
        excel_summary_ws.cell(row=row, column=1).value = 'Check Solver Add-in, select OK'
        row += 2
        excel_summary_ws.cell(row=row, column=1).value = 'Select Data in menu'
        row += 1
        excel_summary_ws.cell(row=row, column=1).value = 'Select Solver in the Analyze section of the ribbon'
        row += 1
        excel_summary_ws.cell(row=row, column=1).value = f'Set "Set Objective" to Total: {total_cell_ref}'
        row += 1
        excel_summary_ws.cell(row=row, column=1).value = f'Set "To:" to Min'
        row += 1
        excel_summary_ws.cell(row=row, column=1).value = f'Set "By Changing Variable Cells:" to the savings plan commits: $B${esp_hourly_commit_first_row}:$B${esp_hourly_commit_last_row},{self._csp_hourly_commit_cell_ref}'
        row += 1
        excel_summary_ws.cell(row=row, column=1).value = f'Set "Select a Solving Method:" to GRG Nonlinear'
        row += 1
        excel_summary_ws.cell(row=row, column=1).value = f'Select Solve'

        # JobStats Worksheet
        max_column_widths = {}
        row = 1
        column = 1
        excel_job_stats_ws.cell(row=row, column=column).value = 'Memory Size (GB)'
        max_column_widths[column] = len(excel_job_stats_ws.cell(row=row, column=column).value)
        column += 1
        for runtime in self.get_ranges(self.runtime_ranges_minutes):
            excel_job_stats_ws.cell(row=row, column=column).value = f'{runtime} Minutes'
            excel_job_stats_ws.merge_cells(start_row=row, start_column=column, end_row=row, end_column=column+2)
            excel_job_stats_ws.cell(row=row, column=column).alignment = XlsAlignment(horizontal='center')
            column += 3
        row += 1
        column = 2
        for runtime in self.get_ranges(self.runtime_ranges_minutes):
            excel_job_stats_ws.cell(row=row, column=column).value = 'Job count'
            max_column_widths[column] = len(excel_job_stats_ws.cell(row=row, column=column).value)
            column += 1
            excel_job_stats_ws.cell(row=row, column=column).value = 'Total duration'
            max_column_widths[column] = len(excel_job_stats_ws.cell(row=row, column=column).value)
            column += 1
            excel_job_stats_ws.cell(row=row, column=column).value = 'Total wait time'
            max_column_widths[column] = len(excel_job_stats_ws.cell(row=row, column=column).value)
            column += 1
        row += 1
        for ram in self.get_ranges(self.ram_ranges_GB):
            column = 1
            excel_job_stats_ws.cell(row=row, column=column).value = f"{ram} GB"
            max_column_widths[column] = max(max_column_widths[column], len(excel_job_stats_ws.cell(row=row, column=column).value))
            column += 1
            for runtime in self.get_ranges(self.runtime_ranges_minutes):
                summary = self.job_data_collector[ram][runtime]
                excel_job_stats_ws.cell(row=row, column=column).value = summary['number_of_jobs']
                excel_job_stats_ws.cell(row=row, column=column+1).value = summary['total_duration_minutes']
                excel_job_stats_ws.cell(row=row, column=column+2).value = summary['total_wait_minutes']
                column += 3
            row += 1
        for column, max_column_width in max_column_widths.items():
            excel_job_stats_ws.column_dimensions[xl_get_column_letter(column)].width = max_column_width + 1
        row += 1
        column = 1
        # Add a chart to show the distribution by job characteristics
        # job_count_chart = BarChart3D()
        # job_count_chart.title = 'Job Count by Duration and Memory Size'
        # job_count_chart.style = 13
        # job_count_chart.y_axis.title = 'Core Hours'
        # job_count_chart.x_axis.title = 'Relative Hour'
        # job_count_chart.width = 30
        # job_count_chart.height = 15
        # excel_core_hours_chart_ws.add_chart(job_count_chart, excel_job_stats_ws.cell(row=row, column=column).coordinate)

        # Config Worksheet
        excel_config_ws.column_dimensions['A'].width = 35
        excel_config_ws.column_dimensions['B'].width = 25
        excel_config_ws.column_dimensions['B'].alignment = XlsAlignment(horizontal='right')
        row = 1
        excel_config_ws[f'A{row}'] = 'Region'
        excel_config_ws[f'B{row}'] = self.config['instance_mapping']['region_name']
        row += 2
        excel_config_ws[f'A{row}'] = 'Minimum CPU Speed (GHz)'
        excel_config_ws[f'B{row}'] = self.config['consumption_model_mapping']['minimum_cpu_speed']
        row += 1
        excel_config_ws[f'A{row}'] = 'Maximum minutes for spot'
        excel_config_ws[f'B{row}'] = self.config['consumption_model_mapping']['maximum_minutes_for_spot']
        row += 2
        esp_term = f"EC2 SP {self.config['consumption_model_mapping']['ec2_savings_plan_duration']}yr {self.config['consumption_model_mapping']['ec2_savings_plan_payment_option']}"
        excel_config_ws[f'A{row}'] = 'EC2 Savings Plan (ESP) Term'
        excel_config_ws[f'B{row}'] = esp_term
        row += 2
        csp_term = f"Compute SP {self.config['consumption_model_mapping']['ec2_savings_plan_duration']}yr {self.config['consumption_model_mapping']['ec2_savings_plan_payment_option']}"
        excel_config_ws[f'A{row}'] = 'Compute Savings Plan (CSP) Term'
        excel_config_ws[f'B{row}'] = csp_term

        # InstanceFamilyRates Worksheet
        instance_info_headings = ['Instance Family', 'OD Rate', 'ESP Rate','ESP Discount', 'ESP Core*Hr Commit', 'CSP Rate', 'CSP Discount', 'CSP Max Core*Hr Commit']
        instance_family_cols = {}
        self._instance_family_col_letters = {}
        for instance_info_heading_column, instance_info_heading in enumerate(instance_info_headings, start=1):
            instance_family_cols[instance_info_heading] = instance_info_heading_column
            self._instance_family_col_letters[instance_info_heading] = xl_get_column_letter(instance_info_heading_column)
            excel_instance_info_ws.cell(row=1, column=instance_info_heading_column, value=instance_info_heading)
            excel_instance_info_ws.column_dimensions[xl_get_column_letter(instance_info_heading_column)].width = len(instance_info_heading) + 1
        self._instance_family_rows = {}
        csp_discounts = {}
        for instance_family_row, instance_family in enumerate(self._instance_families_used, start=2):
            self._instance_family_rows[instance_family] = instance_family_row
            excel_instance_info_ws.cell(row=instance_family_row, column=1, value=instance_family)
            instance_type = self.instance_family_info[instance_family]['MaxInstanceType']
            coreCount = self.instance_type_info[instance_type]['DefaultCores']
            od_rate = self.instance_type_info[instance_type]['pricing']['OnDemand']/coreCount
            excel_instance_info_ws.cell(row=instance_family_row, column=instance_family_cols['OD Rate'], value=od_rate)
            excel_instance_info_ws.cell(row=instance_family_row, column=instance_family_cols['ESP Rate'], value=self.instance_type_info[instance_type]['pricing']['EC2SavingsPlan'][esp_term]/coreCount)
            excel_instance_info_ws.cell(row=instance_family_row, column=instance_family_cols['ESP Discount'], value=f"=({self._instance_family_col_letters['OD Rate']}{instance_family_row}-{self._instance_family_col_letters['ESP Rate']}{instance_family_row})/{self._instance_family_col_letters['OD Rate']}{instance_family_row}")
            excel_instance_info_ws.cell(row=instance_family_row, column=instance_family_cols['ESP Core*Hr Commit'], value=f"={esp_hourly_commit_cell_refs[instance_family]}/{self._instance_family_col_letters['ESP Rate']}{instance_family_row}")
            csp_rate = self.instance_type_info[instance_type]['pricing']['ComputeSavingsPlan'][csp_term]/coreCount
            excel_instance_info_ws.cell(row=instance_family_row, column=instance_family_cols['CSP Rate'], value=csp_rate)
            csp_discounts[instance_family] = (od_rate - csp_rate)/od_rate
            excel_instance_info_ws.cell(row=instance_family_row, column=instance_family_cols['CSP Discount'], value=f"=({self._instance_family_col_letters['OD Rate']}{instance_family_row}-{self._instance_family_col_letters['CSP Rate']}{instance_family_row})/{self._instance_family_col_letters['OD Rate']}{instance_family_row}")
            excel_instance_info_ws.cell(row=instance_family_row, column=instance_family_cols['CSP Max Core*Hr Commit'], value=f"={self._csp_hourly_commit_cell_ref}/{self._instance_family_col_letters['CSP Rate']}{instance_family_row}")

        # CSPs are applied in descending order by size of the discount
        # self._instance_families_by_descending_csp_discounts = sorted(csp_discounts.items(), key=lambda x: x[1], reverse=True)
        self._instance_families_by_descending_csp_discounts = sorted(csp_discounts.items(), key=itemgetter(1), reverse=True)
        logger.debug(f"instance_families_by_descending_csp_discounts: {self._instance_families_by_descending_csp_discounts}")

        # Hourly Worksheet
        self._excel_hourly_ws.freeze_panes = self._excel_hourly_ws['B2']
        hourly_columns = {}
        self._hourly_column_letters = {}
        hourly_field_names = ['Relative Hour', 'Spot Core Hours', 'Spot Costs']
        column = 0
        for field_name in hourly_field_names:
            column += 1
            self._hourly_column_letters[field_name] = xl_get_column_letter(column)
        hourly_instance_family_field_names = [
            'CHr',
            'ESP CHr',  # The actual number of ESP core hours used. Doesn't affect cost calculation, but can be used to get the ESP utilization ration.
            'CSP CHr',  # The actual number of CSP core hours used. This is necessary since the CSP spans instance families which have different discounts.
            'CSP Cost', # CSP cost for this instance family. This is used to get the total amount of the CSP used so far.
            'OD CHr',   # The OD core hours used. Excess core hours not paid for savings plans.
            'OD Cost']  # OD cost
        for instance_family in self._instance_families_used:
            hourly_columns[instance_family] = {}
            self._hourly_column_letters[instance_family] = {}
            for instance_family_field_name in hourly_instance_family_field_names:
                column += 1
                field_name = f"{instance_family} {instance_family_field_name}"
                hourly_field_names.append(field_name)
                hourly_columns[instance_family][field_name] = column
                self._hourly_column_letters[field_name] = xl_get_column_letter(column)
                self._hourly_column_letters[instance_family][instance_family_field_name] = xl_get_column_letter(column)
        hourly_final_field_names = [
            'CSP Cost',   # Total CSP cost. Can be used to calculate CSP utilization
            'OD Cost',    # On demand cost. Don't include ESP and CSP costs because they are fixed per hour
            'Total CHr',    # Total core hours
            'Total OD CHr', # Total on demand core hours. Excludes core hours provided by ESPs and CSPs.
            'Total Cost', # Total cost. Spot, ESP, CSP, and OD
        ]
        for field_name in hourly_final_field_names:
            column += 1
            hourly_field_names.append(field_name)
            self._hourly_column_letters[field_name] = xl_get_column_letter(column)
        excel_hourly_ws_columns = len(hourly_field_names)
        for field_column, field_name in enumerate(hourly_field_names, start=1):
            self._excel_hourly_ws.cell(row=1, column=field_column, value=field_name)
            self._excel_hourly_ws.column_dimensions[xl_get_column_letter(field_column)].width = len(field_name) + 1
        logger.debug(f"excel_hourly_ws_columns: {excel_hourly_ws_columns}")

        # CostSummary Worksheet
        self._total_spot_cell.value = f'=sum(indirect("Hourly!{self._hourly_column_letters["Spot Costs"]}" & {first_hour_cell_ref}+2 & ":{self._hourly_column_letters["Spot Costs"]}" & {last_hour_cell_ref}+2))'
        self._total_esp_cell.value = f'=({last_hour_cell_ref}-{first_hour_cell_ref}+1)*{self._esp_hourly_commit_cell_ref}'
        self._total_csp_cell.value = f'=({last_hour_cell_ref}-{first_hour_cell_ref}+1)*{self._csp_hourly_commit_cell_ref}'
        self._total_od_cell.value = f'=sum(indirect("Hourly!{self._hourly_column_letters["OD Cost"]}" & {first_hour_cell_ref}+2 & ":{self._hourly_column_letters["OD Cost"]}" & {last_hour_cell_ref}+2))'
        total_cell.value =    f'=sum(indirect("Hourly!{self._hourly_column_letters["Total Cost"]}" & {first_hour_cell_ref}+2 & ":{self._hourly_column_letters["Total Cost"]}" & {last_hour_cell_ref}+2))'

        # InstanceFamilySummary Worksheet
        row = 1
        column = 1
        cell = excel_instance_family_summary_ws.cell(row=row, column=column)
        cell.value = 'Instance Family'
        excel_instance_family_summary_ws.column_dimensions[cell.column_letter].width = len(cell.value) + 1
        column += 1
        cell = excel_instance_family_summary_ws.cell(row=row, column=column)
        cell.value = 'Min Core Hours'
        excel_instance_family_summary_ws.column_dimensions[cell.column_letter].width = len(cell.value) + 1
        excel_instance_family_summary_ws.column_dimensions[cell.column_letter].alignment = XlsAlignment(horizontal='right')
        column += 1
        cell = excel_instance_family_summary_ws.cell(row=row, column=column)
        cell.value = 'Avg Core Hours'
        excel_instance_family_summary_ws.column_dimensions[cell.column_letter].width = len(cell.value) + 1
        excel_instance_family_summary_ws.column_dimensions[cell.column_letter].alignment = XlsAlignment(horizontal='right')
        column += 1
        cell = excel_instance_family_summary_ws.cell(row=row, column=column)
        cell.value = 'Max Core Hours'
        excel_instance_family_summary_ws.column_dimensions[cell.column_letter].width = len(cell.value) + 1
        excel_instance_family_summary_ws.column_dimensions[cell.column_letter].alignment = XlsAlignment(horizontal='right')
        column += 1
        cell = excel_instance_family_summary_ws.cell(row=row, column=column)
        cell.value = 'Min OD Cost'
        excel_instance_family_summary_ws.column_dimensions[cell.column_letter].width = len(cell.value) + 1
        excel_instance_family_summary_ws.column_dimensions[cell.column_letter].alignment = XlsAlignment(horizontal='right')
        column += 1
        cell = excel_instance_family_summary_ws.cell(row=row, column=column)
        cell.value = 'Avg OD Cost'
        excel_instance_family_summary_ws.column_dimensions[cell.column_letter].width = len(cell.value) + 1
        excel_instance_family_summary_ws.column_dimensions[cell.column_letter].alignment = XlsAlignment(horizontal='right')
        column += 1
        cell = excel_instance_family_summary_ws.cell(row=row, column=column)
        cell.value = 'Max OD Cost'
        excel_instance_family_summary_ws.column_dimensions[cell.column_letter].width = len(cell.value) + 1
        excel_instance_family_summary_ws.column_dimensions[cell.column_letter].alignment = XlsAlignment(horizontal='right')

        instance_family_first_row = row + 1
        for instance_family in self._instance_families_used:
            row += 1
            column = 1
            excel_instance_family_summary_ws.cell(row=row, column=column).value = f'{instance_family}'

            column += 1
            cell = excel_instance_family_summary_ws.cell(row=row, column=column)
            cell.value = f'=min(indirect("Hourly!{self._hourly_column_letters[instance_family + " CHr"]}" & CostSummary!{first_hour_cell.coordinate}+2 & ":{self._hourly_column_letters[instance_family + " CHr"]}" & CostSummary!{last_hour_cell.coordinate}+2))'
            column += 1
            cell = excel_instance_family_summary_ws.cell(row=row, column=column)
            cell.value = f'=average(indirect("Hourly!{self._hourly_column_letters[instance_family + " CHr"]}" & CostSummary!{first_hour_cell.coordinate}+2 & ":{self._hourly_column_letters[instance_family + " CHr"]}" & CostSummary!{last_hour_cell.coordinate}+2))'
            column += 1
            cell = excel_instance_family_summary_ws.cell(row=row, column=column)
            cell.value = f'=max(indirect("Hourly!{self._hourly_column_letters[instance_family + " CHr"]}" & CostSummary!{first_hour_cell.coordinate}+2 & ":{self._hourly_column_letters[instance_family + " CHr"]}" & CostSummary!{last_hour_cell.coordinate}+2))'

            column += 1
            cell = excel_instance_family_summary_ws.cell(row=row, column=column)
            cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
            cell.value = f'=min(indirect("Hourly!{self._hourly_column_letters[instance_family + " OD Cost"]}" & CostSummary!{first_hour_cell.coordinate}+2 & ":{self._hourly_column_letters[instance_family + " OD Cost"]}" & CostSummary!{last_hour_cell.coordinate}+2))'
            column += 1
            cell = excel_instance_family_summary_ws.cell(row=row, column=column)
            cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
            cell.value = f'=average(indirect("Hourly!{self._hourly_column_letters[instance_family + " OD Cost"]}" & CostSummary!{first_hour_cell.coordinate}+2 & ":{self._hourly_column_letters[instance_family + " OD Cost"]}" & CostSummary!{last_hour_cell.coordinate}+2))'
            column += 1
            cell = excel_instance_family_summary_ws.cell(row=row, column=column)
            cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
            cell.value = f'=max(indirect("Hourly!{self._hourly_column_letters[instance_family + " OD Cost"]}" & CostSummary!{first_hour_cell.coordinate}+2 & ":{self._hourly_column_letters[instance_family + " OD Cost"]}" & CostSummary!{last_hour_cell.coordinate}+2))'
        instance_family_last_row = row
        row += 1
        excel_instance_family_summary_ws.cell(row=row, column=1).value = 'Total'
        for column in range(2, 8):
            cell = excel_instance_family_summary_ws.cell(row=row, column=column)
            if self._instance_families_used:
                cell.value = f"=sum({xl_get_column_letter(column)}{instance_family_first_row}:{xl_get_column_letter(column)}{instance_family_last_row})"
            else:
                cell.value = 0
            if column > 4:
                cell.number_format = FORMAT_CURRENCY_USD_SIMPLE

        # Core Hours Charts

        column = 1
        row = 1

        # Stacked Core Hours Chart
        # Stacked line chart with the number of core hours per instance family
        core_hours_chart = XlLineChart()
        core_hours_chart.title = 'Core Hours by Instance Family'
        core_hours_chart.style = 13
        core_hours_chart.y_axis.title = 'Core Hours'
        core_hours_chart.x_axis.title = 'Relative Hour'
        core_hours_chart.grouping = 'stacked'
        core_hours_chart.width = 30
        core_hours_chart.height = 15
        for instance_family in self._instance_families_used:
            column = hourly_columns[instance_family][f'{instance_family} CHr']
            data_series = XlReference(self._excel_hourly_ws, min_col=column, min_row=1, max_col=column, max_row=last_hour_cell.value + 2)
            core_hours_chart.add_data(data_series, titles_from_data=True)
        excel_core_hours_chart_ws.add_chart(core_hours_chart, 'A1')
        row += 30

        # Core Hours Chart by instance family
        for instance_family in self._instance_families_used:
            core_hours_chart = XlLineChart()
            core_hours_chart.title = f'{instance_family} Core Hours'
            core_hours_chart.style = 13
            core_hours_chart.y_axis.title = 'Core Hours'
            core_hours_chart.x_axis.title = 'Relative Hour'
            core_hours_chart.width = 30
            core_hours_chart.height = 15
            column = hourly_columns[instance_family][f'{instance_family} CHr']
            data_series = XlReference(self._excel_hourly_ws, min_col=column, min_row=1, max_col=column, max_row=last_hour_cell.value + 2)
            core_hours_chart.add_data(data_series, titles_from_data=True)
            cell = excel_core_hours_chart_ws.cell(row=row, column=1)
            excel_core_hours_chart_ws.add_chart(core_hours_chart, cell.coordinate)
            row += 30

    def _write_excel_row(self) -> None:
        '''
        Write hourly stats to Excel file
        '''
        row = self._relative_hour + 2
        logger.debug(f"row: {row}")
        self._excel_hourly_ws.cell(row=row, column=1, value=self._relative_hour)
        self._excel_hourly_ws.cell(row=row, column=2, value=self._hourly_costs['spot']['core hours'])
        self._excel_hourly_ws.cell(row=row, column=3, value=self._hourly_costs['spot']['total costs'])
        od_cost_formula = '=0'
        csp_cost_total_formula = '0'
        total_core_hour_formula = '=0'
        total_od_core_hour_formula = '=0'
        for instance_family, instance_family_csp_discount in self._instance_families_by_descending_csp_discounts:
            instance_family_row = self._instance_family_rows[instance_family]
            # Total core hours
            core_hours = self._hourly_costs[ComputeInstance.ON_DEMAND][instance_family]['core hours']
            self._excel_hourly_ws[f"{self._hourly_column_letters[instance_family]['CHr']}{row}"] = core_hours
            # ESP core hours actually used
            self._excel_hourly_ws[f"{self._hourly_column_letters[instance_family]['ESP CHr']}{row}"] = f"=min({self._hourly_column_letters[instance_family]['CHr']}{row}, InstanceFamilyRates!${self._instance_family_col_letters['ESP Core*Hr Commit']}${instance_family_row})"
            # CSP core hours used by this instance family
            # First calculate the remaining instance family core hours by subtracting the ESP core hours.
            # First calculate the remaining CSP commit available.
            # Then use the available CSP dollars to calculate the number of CSP core hours available
            # Then use as many of those CSP core hours as possible.
            self._excel_hourly_ws[f"{self._hourly_column_letters[instance_family]['CSP CHr']}{row}"] = f"=min({self._hourly_column_letters[instance_family]['CHr']}{row}-{self._hourly_column_letters[instance_family]['ESP CHr']}{row}, ({self._csp_hourly_commit_cell_ref}-({csp_cost_total_formula}))/InstanceFamilyRates!${self._instance_family_col_letters['CSP Rate']}${instance_family_row})"
            # CSP Cost
            self._excel_hourly_ws[f"{self._hourly_column_letters[instance_family]['CSP Cost']}{row}"] = f"={self._hourly_column_letters[instance_family]['CSP CHr']}{row}*InstanceFamilyRates!${self._instance_family_col_letters['CSP Rate']}${instance_family_row}"
            # OD core hours
            self._excel_hourly_ws[f"{self._hourly_column_letters[instance_family]['OD CHr']}{row}"] = f"={self._hourly_column_letters[instance_family]['CHr']}{row}-{self._hourly_column_letters[instance_family]['ESP CHr']}{row}-{self._hourly_column_letters[instance_family]['CSP CHr']}{row}"
            self._excel_hourly_ws[f"{self._hourly_column_letters[instance_family]['OD Cost']}{row}"] = f"={self._hourly_column_letters[instance_family]['CSP CHr']}{row}*InstanceFamilyRates!$C${instance_family_row}+{self._hourly_column_letters[instance_family]['OD CHr']}{row}*InstanceFamilyRates!$B${instance_family_row}"
            csp_cost_total_formula += f"+{self._hourly_column_letters[instance_family]['CSP Cost']}{row}"
            od_cost_formula += f"+{self._hourly_column_letters[instance_family]['OD Cost']}{row}"
            total_core_hour_formula += f"+{self._hourly_column_letters[instance_family]['CHr']}{row}"
            total_od_core_hour_formula += f"+{self._hourly_column_letters[instance_family]['OD CHr']}{row}"
        self._excel_hourly_ws[f"{self._hourly_column_letters['CSP Cost']}{row}"] = f"={csp_cost_total_formula}"
        self._excel_hourly_ws[f"{self._hourly_column_letters['OD Cost']}{row}"] = f"{od_cost_formula}"
        self._excel_hourly_ws[f"{self._hourly_column_letters['Total CHr']}{row}"] = f"{total_core_hour_formula}"
        self._excel_hourly_ws[f"{self._hourly_column_letters['Total OD CHr']}{row}"] = f"{total_od_core_hour_formula}"
        self._excel_hourly_ws[f"{self._hourly_column_letters['Total Cost']}{row}"] = f"={self._hourly_column_letters['Spot Costs']}{row}+{self._esp_hourly_commit_cell_ref}+{self._csp_hourly_commit_cell_ref}+{self._hourly_column_letters['OD Cost']}{row}"

    def _finish_excel(self):
        self._excel_wb.save(self._excel_filename)

    def _process_scratch_csv(self) -> None:
        while self._read_scratch_csv():
            self._write_csv_row()
            self._write_excel_row()

def main():
    try:
        parser = argparse.ArgumentParser(description="Model jobs running on an ideal computer cluster.", formatter_class=argparse.ArgumentDefaultsHelpFormatter)
        jobs_csv_group = parser.add_mutually_exclusive_group(required=True)
        jobs_csv_group.add_argument("--unsorted-jobs-csv", help="CSV file with parsed job info from scheduler parser that hasn't been sorted by eligible_time.")
        jobs_csv_group.add_argument("--sorted-jobs-csv", help="CSV file with parsed job info from scheduler parser that has been sorted by eligible_time.")
        parser.add_argument("--starttime", help="Select jobs after the specified time. Format YYYY-MM-DDTHH:MM:SS")
        parser.add_argument("--endtime", help="Select jobs before the specified time. Format YYYY-MM-DDTHH:MM:SS")
        parser.add_argument("--config", required=False, default=f'{dirname(__file__)}/config.yml', help="Configuration file.")
        parser.add_argument("--acknowledge-config", required=False, action='store_const', const=True, default=False, help="Acknowledge configuration file contents so don't get prompt.")
        parser.add_argument("--scheduling-frequency", required=False, default=None, help="Number of minutes between when jobs are scheduled in minutes. Overrides value in config file.")
        parser.add_argument("--share-instances", type=bool, required=False, default=None, help="Run multiple jobs per instance. Overrides value in config file.")
        parser.add_argument("--output-dir", required=False, default="output", help="Directory where output will be written")
        parser.add_argument("--queues", default=None, help="Comma separated list of regular expressions of queues to include/exclude. Prefix the queue with '-' to exclude. The regular expressions are evaluated in the order given and the first match has precedence and stops further evaluations. Regular expressions have an implicit ^ at the beginning.")
        parser.add_argument("--projects", default=None, help="Comma separated list of regular expressions of projects to include/exclude. Prefix the project with '-' to exclude. The regular expressions are evaluated in the order given and the first match has precedence and stops further evaluations. Regular expressions have an implicit ^ at the beginning.")
        parser.add_argument("--disable-version-check", action='store_const', const=True, default=False, help="Disable git version check")
        parser.add_argument("--debug", '-d', action='count', default=0, help="Enable debug mode")
        args = parser.parse_args()

        # Configure logfile
        timestamp_str = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        output_dir = realpath(args.output_dir)
        if not path.exists(output_dir):
            logger.info(f"Output directory ({output_dir}) doesn't exist, creating")
            makedirs(output_dir)
        log_file_name = path.join(output_dir, f"ModelComputeCluster-{timestamp_str}.log")
        logger_FileHandler = logging.FileHandler(filename=log_file_name)
        logger_FileHandler.setFormatter(logger_formatter)
        logger.addHandler(logger_FileHandler)

        if args.debug:
            debug_level = args.debug
            debug_print_event_summary = debug_level > 1
            debug_print_event_details = debug_level > 2
            debug_insert_instance = debug_level > 3
            debug_remove_instance = debug_level > 3
            logger.setLevel(logging.DEBUG)
            CSVLogParser_logger.setLevel(logging.DEBUG)
            SchedulerJobInfo_logger.setLevel(logging.DEBUG)
            SchedulerLogParser_logger.setLevel(logging.DEBUG)
            VersionCheck_logger.setLevel(logging.DEBUG)

        if not args.disable_version_check and not VersionCheck().check_git_version():
            exit(1)

        if args.unsorted_jobs_csv:
            args.sorted_jobs_csv = path.join(args.output_dir, 'sorted_jobs.csv')
            logger.info(f"Sorting {args.unsorted_jobs_csv} by eligible_time into {args.sorted_jobs_csv}")
            job_sorter = JobSorter(args.unsorted_jobs_csv, args.sorted_jobs_csv)
            job_sorter.sort_jobs()

        logger.info(f"Reading job data from {args.sorted_jobs_csv}")
        csv_parser = CSVLogParser(args.sorted_jobs_csv, None, args.starttime, args.endtime)

        logger.info('Started job modelling')

        compute_cluster_model = ComputeClusterModel(csv_parser, args.config, args.output_dir, args.starttime, args.endtime, queue_filters=args.queues, project_filters=args.projects)

        # Print out configuration information
        logger.info(f"""Configuration:
        {json.dumps(compute_cluster_model.config, indent=4)}""")
        acknowledge_config = args.acknowledge_config
        while not acknowledge_config:
            print(f"\nIs the correct configuration? (y/n) ")
            answer = input().lower()
            if answer == 'n':
                exit(1)
            elif answer == 'y':
                acknowledge_config = True

        compute_cluster_model.schedule_jobs()
    except Exception as e:
        logger.exception(f"Unhandled exception")
        exit(1)

if __name__ == '__main__':
    main()
