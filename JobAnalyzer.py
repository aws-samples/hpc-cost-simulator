#!/usr/bin/env python3
'''
Analyze the jobs parsed from scheduler logs

This module holds all the supporting functions required to initialize the data structure required to parse scheduler log files
from all schedulers. It builds the list of instances by memory / physical cores / core speed and gets the pricing for them to
allow jobs to be mapped to instance types

It also holds the function to map a job to the right instance type.

Copyright Amazon.com, Inc. or its affiliates. All Rights Reserved.
SPDX-License-Identifier: MIT-0
'''

from AcceleratorLogParser import AcceleratorLogParser, logger as AcceleratorLogParser_logger
import argparse
from botocore.exceptions import ClientError, NoCredentialsError
from colored import fg
from JobAnalyzerBase import JobAnalyzerBase
from config_schema import check_schema
from copy import deepcopy
import csv
from CSVLogParser import CSVLogParser, logger as CSVLogParser_logger
from datetime import datetime, time, timedelta
from EC2InstanceTypeInfoPkg.EC2InstanceTypeInfo import EC2InstanceTypeInfo
import json
from LSFLogParser import LSFLogParser, logger as LSFLogParser_logger
import logging
from math import ceil
from openpyxl import Workbook as XlsWorkbook
from openpyxl.chart import BarChart3D, LineChart as XlLineChart, Reference as XlReference
from openpyxl.styles import Alignment as XlsAlignment, Protection as XlsProtection
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE
from openpyxl.utils import get_column_letter as xl_get_column_letter
import operator
from os import listdir, makedirs, path, remove
from os.path import dirname, realpath
import re
from SchedulerJobInfo import logger as SchedulerJobInfo_logger, SchedulerJobInfo
from SchedulerLogParser import SchedulerLogParser, logger as SchedulerLogParser_logger
from SlurmLogParser import SlurmLogParser, logger as SlurmLogParser_logger
from sys import exit
from VersionCheck import logger as VersionCheck_logger, VersionCheck

logger = logging.getLogger(__file__)
logger_formatter = logging.Formatter('%(levelname)s:%(asctime)s: %(message)s')
logger_streamHandler = logging.StreamHandler()
logger_streamHandler.setFormatter(logger_formatter)
logger.addHandler(logger_streamHandler)
logger.propagate = False
logger.setLevel(logging.INFO)

SECONDS_PER_MINUTE = 60
MINUTES_PER_HOUR = 60
SECONDS_PER_HOUR = SECONDS_PER_MINUTE * MINUTES_PER_HOUR

class JobCost:
    def __init__(self, job: SchedulerJobInfo, spot: bool, instance_family: str, instance_type: str, rate: float):
        self.job = job
        self.spot = spot
        self.instance_family = instance_family
        self.instance_type = instance_type
        self.rate = rate

class JobAnalyzer(JobAnalyzerBase):

    def __init__(self, scheduler_parser: SchedulerLogParser, config_filename: str, output_dir: str, starttime: str, endtime: str, queue_filters: str, project_filters: str) -> None:
        '''
        Constructor

        Args:
            scheduler_parser (SchedulerLogParser): Job parser
            config_filename (str): Configuration file
            output_dir (str): Output directory
            starttime (str): Select jobs after the specified time
            endtime (str): Select jobs after the specified time
            queue_filters (str): Queue filters
            project_filters (str): Project filters
        Returns:
            None
        '''
        super().__init__(scheduler_parser, config_filename, output_dir, starttime, endtime, queue_filters, project_filters)

        self._scheduler_parser = scheduler_parser
        self._config_filename = config_filename
        self._output_dir = realpath(output_dir)

    def get_hourly_files(self):
        '''
        Gets the hourly output files for the current job

        Input: none

        Output:
            Sorted list of output filenames
        '''
        all_files = listdir(self._output_dir)
        output_files = []
        prefix = path.basename("hourly-")
        for file in all_files:
            if file.startswith(prefix) and file[-4:] == ".csv":
                output_file = f"{self._output_dir}/" + file
                output_files.append(output_file)
        output_files.sort()
        return output_files

    def _cleanup_hourly_files(self):
        '''
        Delete the hourly output files so old results aren't included in new run.

        Input: none

        Output:
            Sorted list of output filenames
        '''
        hourly_files = self.get_hourly_files()
        for hourly_file in hourly_files:
            remove(hourly_file)

    def _update_hourly_stats(self, round_hour: int, minutes_within_hour: float, core_hours: float, total_cost_per_hour: float, spot: bool, instance_family: str) -> None:
        '''
        Update the hourly stats dict with a portion of the cost of a job that fits within a round hour.

        A single job's cost may complete within the same round hour or span beyond it to multiple hours.
        Jobs are broken down by the Spot threshold.

        Args:
            round_hour (int): the hour of the HH:00:00 start
            minutes_within_hour (float): number of minutes the job ran within that round hour
            core_hours: Number of core hours
            total_cost_per_hour (float): the total cost of all instances used to run the job if they ran for a full hour.
            spot (bool): True if spot instance
            instance_family (str): Instance family used for the job
        '''
        round_hour = int(round_hour)
        if self._starttime:
            if round_hour * SECONDS_PER_HOUR < self._starttime_dt.timestamp():
                logger.debug(f"Skipping round_hour={round_hour} timestamp={round_hour * SECONDS_PER_HOUR} {datetime.fromtimestamp(round_hour * SECONDS_PER_HOUR)}")
                logger.debug(f"    starttime  hour={int(self._starttime_dt.timestamp() / SECONDS_PER_HOUR)} timestamp={self._starttime_dt.timestamp()} {self._starttime_dt}")
                logger.debug(f"    endtime    hour={int(self._endtime_dt.timestamp() / SECONDS_PER_HOUR)} timestamp={self._endtime_dt.timestamp()} {self._endtime_dt}")
                return

        if self._endtime:
            if round_hour * SECONDS_PER_HOUR > self._endtime_dt.timestamp():
                logger.debug(f"Skipping round_hour={round_hour} timestamp={round_hour * SECONDS_PER_HOUR} {datetime.fromtimestamp(round_hour * SECONDS_PER_HOUR)}")
                logger.debug(f"    starttime  hour={int(self._starttime_dt.timestamp() / SECONDS_PER_HOUR)} timestamp={self._starttime_dt.timestamp()} {self._starttime_dt}")
                logger.debug(f"    endtime    hour={int(self._endtime_dt.timestamp() / SECONDS_PER_HOUR)} timestamp={self._endtime_dt.timestamp()} {self._endtime_dt}")
                return

        if round_hour not in self.hourly_stats:
            self._init_hourly_stats_hour(round_hour)
        purchase_option = 'spot' if spot == True else 'on_demand'
        cost = minutes_within_hour / 60 * total_cost_per_hour
        if spot:
            self.hourly_stats[round_hour][purchase_option] += cost
            self.total_stats[purchase_option] += cost
        else:
            self.hourly_stats[round_hour][purchase_option]['total'] += cost
            self.hourly_stats[round_hour][purchase_option][instance_family] = cost + self.hourly_stats[round_hour][purchase_option].get(instance_family, 0)
            self.hourly_stats[round_hour][purchase_option]['core_hours'][instance_family] = self.hourly_stats[round_hour][purchase_option]['core_hours'].get(instance_family, 0) + core_hours
            self.total_stats[purchase_option]['total'] += cost
            self.total_stats[purchase_option]['instance_families'][instance_family] = cost + self.total_stats[purchase_option]['instance_families'].get(instance_family, 0)

    def _process_hourly_jobs(self) -> None:
        '''
        Process hourly job CSV files

        Sequentially processes the hourly output files to build an hourly-level cost simulation
        '''
        if not self.instance_family_info:
            self.get_instance_type_info()

        logger.info('')
        logger.info(f"Post processing hourly jobs data:\n")

        hourly_files = self.get_hourly_files()
        if len(hourly_files) == 0:
            logger.warning(f"No hourly jobs files found")
            exit(0)
        logger.info(f"Found {len(hourly_files)} hourly jobs files.")

        for hourly_file_index, hourly_file in enumerate(hourly_files):
            logger.debug(f"Processing {hourly_file}")
            with open(hourly_file, 'r') as hourly_job_file_fh:
                csv_reader = csv.reader(hourly_job_file_fh, dialect='excel')
                field_names = next(csv_reader)
                num_jobs = 0
                line_number = 1
                while True:
                    try:
                        job_field_values_array = next(csv_reader)
                    except StopIteration:
                        break
                    num_jobs += 1
                    line_number += 1
                    job_field_values = {}
                    for i, field_name in enumerate(field_names):
                        job_field_values[field_name] = job_field_values_array[i]
                    start_time = SchedulerJobInfo.str_to_datetime(job_field_values['start_time']).timestamp()
                    job_id = job_field_values['Job id']
                    num_hosts = int(job_field_values['Num Hosts'])
                    job_runtime_minutes = float(job_field_values['Runtime (minutes)'])
                    instance_type = job_field_values['Instance type']
                    instance_family = job_field_values['Instance Family']
                    spot_eligible = job_field_values['Spot'] == 'True'
                    on_demand_rate = float(job_field_values['Hourly Rate'])
                    total_on_demand_cost = job_field_values['Total Cost']

                    end_time = start_time + job_runtime_minutes * 60
                    total_hourly_rate = on_demand_rate * num_hosts
                    num_cores = self.instance_type_info[instance_type]['CoreCount']

                    logger.debug(f"    job {job_id}: line {line_number}")
                    logger.debug(f"        num_hosts={num_hosts}")
                    logger.debug(f"        start_time={start_time}")
                    logger.debug(f"        start_time={start_time}")
                    logger.debug(f"        end_time  ={end_time}")
                    logger.debug(f"        instance_family={instance_family}")
                    logger.debug(f"        total cost={total_on_demand_cost}")
                    logger.debug(f"        spot_eligible={spot_eligible}")
                    logger.debug(f"        job_runtime_minutes={job_runtime_minutes}")
                    logger.debug(f"        total_hourly_rate={total_hourly_rate}")
                    logger.debug(f"        num_cores={num_cores}")

                    if spot_eligible:
                        purchase_option = 'spot'
                    else:
                        purchase_option = 'on_demand'
                    self._instance_types_used[purchase_option][instance_type] = self._instance_types_used[purchase_option].get(instance_type, 0) + num_hosts
                    self._instance_families_used[purchase_option][instance_family] = self._instance_families_used[purchase_option].get(instance_family, 0) + num_hosts

                    round_hour = int(start_time//SECONDS_PER_HOUR)
                    round_hour_seconds = round_hour * SECONDS_PER_HOUR
                    logger.debug(f"        round_hour: {round_hour}")
                    logger.debug(f"        round_hour_seconds: {round_hour_seconds}")
                    while round_hour_seconds < end_time:
                        next_round_hour = round_hour + 1
                        next_round_hour_seconds = next_round_hour * SECONDS_PER_HOUR
                        if round_hour_seconds <= start_time < next_round_hour_seconds:
                            logger.debug(f"        Job started in this hour")
                            if end_time <= next_round_hour_seconds:
                                logger.debug(f"        job ended within hour")
                                runtime_minutes = (end_time - start_time)/60
                            else:
                                logger.debug(f"        job spills into the next hour")
                                runtime_minutes = (next_round_hour_seconds - start_time)/60
                        elif start_time <= round_hour_seconds and end_time > next_round_hour_seconds:
                            logger.debug(f"        Job started before this hour and runs throughout the hour")
                            runtime_minutes = 60
                        elif start_time < round_hour_seconds and end_time <= next_round_hour_seconds:
                            logger.debug(f"        Job started in prev hour, ends in this one")
                            runtime_minutes = (end_time - round_hour_seconds)/60
                        else:
                            logger.error(f"{file}, line {line_number}: Record failed to process correctly: {','.join(job_field_values_array)}")
                        core_hours = runtime_minutes * num_hosts * num_cores / 60
                        self._update_hourly_stats(round_hour, runtime_minutes, core_hours, total_hourly_rate, spot_eligible, instance_family)
                        round_hour += 1
                        round_hour_seconds = round_hour * SECONDS_PER_HOUR
            logger.debug(f"    Finished processing ({num_jobs} jobs)")
	    # Print a progress message for every 24 hours of data
            if hourly_file_index and (hourly_file_index % 24 == 0):
                logger.info(f"Processed {hourly_file_index + 1} / {len(hourly_files)} ({round((hourly_file_index + 1) / len(hourly_files) * 100)} %) of hourly job files.")

    def _add_job_to_hourly_bucket(self, job_cost_data: JobCost) -> None:
        '''
        Put job into an hourly bucket

        The hourly buckets get written into files for scalability, but for performance reasons they are only
        written when the bucket contains a configurable number of jobs.
        This prevents a file open, write, close for each job.
        '''
        job = job_cost_data.job
        round_hour = int(job.start_time_dt.timestamp()//SECONDS_PER_HOUR)
        if round_hour not in self.jobs_by_hours:
            self.jobs_by_hours[round_hour] = [job_cost_data]
        else:
            self.jobs_by_hours[round_hour].append(job_cost_data)
        self._hourly_jobs_to_be_written +=1
        if self._hourly_jobs_to_be_written >= self.config['consumption_model_mapping']['job_file_batch_size']:
            self._write_hourly_jobs_buckets_to_file()

    def _write_hourly_jobs_buckets_to_file(self) -> None:
        '''
        Write hourly jobs to files

        This is done in batches to reduce the number of file opens and closes.
        '''
        for round_hour, jobs in self.jobs_by_hours.items():
            hourly_file_name = path.join(self._output_dir, f"hourly-{round_hour}.csv")
            with open(hourly_file_name, 'a+') as job_file:
                if job_file.tell() == 0:    # Empty file - add headers
                    job_file.write('start_time,Job id,Num Hosts,Runtime (minutes),memory (GB),Wait time (minutes),Instance type,Instance Family,Spot,Hourly Rate,Total Cost\n')
                for job_cost_data in jobs:
                    job = job_cost_data.job
                    runtime_minutes = round(job.run_time_td.total_seconds()/60, 4)
                    runtime_hours = runtime_minutes / 60
                    total_on_demand_cost = round(job.num_hosts * runtime_hours * job_cost_data.rate, 6)
                    wait_time_minutes = round(job.wait_time_td.total_seconds()/60, 4)
                    job_file.write(f"{SchedulerJobInfo.datetime_to_str(job.start_time_dt)},{job.job_id},{job.num_hosts},{runtime_minutes},{job.max_mem_gb},{wait_time_minutes},{job_cost_data.instance_type},{job_cost_data.instance_family},{job_cost_data.spot},{job_cost_data.rate},{total_on_demand_cost}\n")
        self._hourly_jobs_to_be_written = 0
        self.jobs_by_hours = {}

    def _write_hourly_stats(self) -> None:
        '''
        Write hourly stats to CSV and Excel files
        '''
        self._write_hourly_stats_csv()

        self._write_hourly_stats_xlsx()

    def _write_hourly_stats_csv(self):
        '''
        Write hourly stats to CSV file
        '''
        hourly_stats_csv = path.join(self._output_dir, 'hourly_stats.csv')
        logger.info('')
        logger.info(f"Writing hourly stats to {hourly_stats_csv}")

        hour_list = list(self.hourly_stats.keys())
        hour_list.sort()
        logger.debug(f"{len(hour_list)} hours in hourly_stats")
        number_of_hours = 0
        with open(hourly_stats_csv, 'w+') as hourly_stats_fh:
            # convert from absolute hour to relative one (obfuscation)
            csv_writer = csv.writer(hourly_stats_fh, dialect='excel')
            instance_families = sorted(self._instance_families_used['on_demand'].keys())
            field_names = ['Relative Hour','Total OnDemand Costs','Total Spot Costs'] + instance_families
            csv_writer.writerow(field_names)
            if hour_list:
                first_hour = hour_list[0]
                last_hour = hour_list[-1]
                logger.info(f"First hour = {first_hour} = {datetime.fromtimestamp(first_hour * SECONDS_PER_HOUR)}")
                logger.info(f"Last  hour = {last_hour} = {datetime.fromtimestamp(last_hour * SECONDS_PER_HOUR)}")
                logger.info(f"{last_hour - first_hour + 1} total hours")
                prev_relative_hour = 0
                for hour in hour_list:
                    relative_hour = hour - first_hour
                    while prev_relative_hour < relative_hour - 1:   # add zero values for all missing hours
                        field_values = [prev_relative_hour + 1, 0, 0]
                        for instance_family in instance_families:
                            field_values.append(0)
                        csv_writer.writerow(field_values)
                        logger.debug(f"    empty hour: {field_values}")
                        number_of_hours += 1
                        prev_relative_hour += 1
                    field_values = [relative_hour, round(self.hourly_stats[hour]['on_demand']['total'], 6), round(self.hourly_stats[hour]['spot'], 6)]
                    for instance_family in instance_families:
                        field_values.append(round(self.hourly_stats[hour]['on_demand'].get(instance_family, 0), 6))
                    csv_writer.writerow(field_values)

                    number_of_hours += 1
                    prev_relative_hour = relative_hour

        summary_stats_csv = path.join(self._output_dir, 'summary_stats.csv')
        logger.info('')
        logger.info(f"Writing summary stats to {summary_stats_csv}")

        with open (summary_stats_csv, 'w+') as summary_stats_fh:
            csv_writer = csv.writer(summary_stats_fh, dialect='excel')
            instance_families = sorted(self._instance_families_used['on_demand'].keys())
            field_names = ['', 'OnDemand Costs','Spot Costs']
            for instance_family in instance_families:
                field_names.append(f"{instance_family} OD Costs")
            csv_writer.writerow(field_names)

            field_values = ['Total', round(self.total_stats['on_demand']['total'], 6), round(self.total_stats['spot'], 6)]
            for instance_family in instance_families:
                field_values.append(round(self.total_stats['on_demand']['instance_families'].get(instance_family, 0), 6))
            csv_writer.writerow(field_values)

            field_values = ['Hourly average', round(self.total_stats['on_demand']['total']/number_of_hours, 6), round(self.total_stats['spot']/number_of_hours, 6)]
            for instance_family in instance_families:
                field_values.append(round(self.total_stats['on_demand']['instance_families'].get(instance_family, 0)/number_of_hours, 6))
            csv_writer.writerow(field_values)

            field_values = ['Annual average', round(self.total_stats['on_demand']['total']/number_of_hours, 6), round(self.total_stats['spot']/number_of_hours, 6)]
            for instance_family in instance_families:
                field_values.append(round(self.total_stats['on_demand']['instance_families'].get(instance_family, 0)/number_of_hours, 6))
            csv_writer.writerow(field_values)

            field_values = ['Monthly average']

    def parse_hourly_stats_csv(self, hourly_stats_csv: str) -> None:
        logger.info(f"Parsing {hourly_stats_csv}")
        if not path.exists(hourly_stats_csv):
            logger.error(f"{hourly_stats_csv} doesn't exist")
            exit(2)
        if not self.instance_family_info:
            self.get_instance_type_info()
        self._clear_job_stats()
        self.hourly_stats = {}
        with open(hourly_stats_csv, 'r') as hourly_stats_csv_fh:
            csv_reader = csv.reader(hourly_stats_csv_fh, dialect='excel')
            hourly_stats_field_names = next(csv_reader)
            logger.info(f"hourly_stats_field_names: {hourly_stats_field_names}")
            num_hours = 0
            while True:
                try:
                    hourly_stats_values = next(csv_reader)
                except StopIteration:
                    break
                num_hours += 1
                instance_family_on_demand_costs = {}
                for field_index, field_name in enumerate(hourly_stats_field_names):
                    field_value = hourly_stats_values[field_index]
                    logger.debug(f"{field_name}: {field_value}")
                    if field_name == 'Relative Hour':
                        relative_hour = int(field_value)
                    elif field_name == 'Total OnDemand Costs':
                        if field_value == '0':
                            field_value = int(field_value)
                        else:
                            field_value = float(field_value)
                        total_ondemand_costs = field_value
                    elif field_name == 'Total Spot Costs':
                        if field_value == '0':
                            field_value = int(field_value)
                        else:
                            field_value = float(field_value)
                        total_spot_costs = field_value
                    elif field_name in self.instance_family_info:
                        instance_family = field_name
                        if field_value == '0':
                            on_demand_costs = int(field_value)
                        else:
                            on_demand_costs = float(field_value)
                        instance_family_on_demand_costs[instance_family] = on_demand_costs
                    else:
                        raise ValueError(f"Unknown field name in {hourly_stats_csv}: {field_name}")

                self._init_hourly_stats_hour(relative_hour)
                for instance_family, instance_family_on_demand_cost in instance_family_on_demand_costs.items():
                    instance_type = self.instance_family_info[instance_family]['MaxInstanceType']
                    coreCount = self.instance_type_info[instance_type]['CoreCount']
                    od_rate = self.instance_type_info[instance_type]['pricing']['OnDemand']/coreCount
                    core_hours = instance_family_on_demand_cost / od_rate
                    self.hourly_stats[relative_hour]['on_demand'][instance_family] = instance_family_on_demand_costs[instance_family]
                    self.hourly_stats[relative_hour]['on_demand']['core_hours'][instance_family] = core_hours
                    self._instance_families_used['on_demand'][instance_family] = 1
                    self.total_stats['on_demand']['instance_families'][instance_family] = instance_family_on_demand_cost
                self.total_stats['on_demand']['total'] += total_ondemand_costs
                self.total_stats['spot'] += total_spot_costs

    def _write_hourly_stats_xlsx(self):
        '''
        Write hourly stats to Excel file
        '''
        hourly_stats_xlsx = path.join(self._output_dir, 'hourly_stats.xlsx')
        logger.info('')
        logger.info(f"Writing hourly stats to {hourly_stats_xlsx}")

        instance_families = sorted(self._instance_families_used['on_demand'].keys())

        excel_wb = XlsWorkbook()
        xls_locked = XlsProtection(locked=True)
        xls_unlocked = XlsProtection(locked=False)

        excel_wb.calcMode = 'auto'

        # Create worksheets
        excel_summary_ws = excel_wb.active
        excel_summary_ws.title = 'CostSummary'
        #excel_summary_ws.protection.sheet = xls_locked

        excel_job_stats_ws = excel_wb.create_sheet(title='JobStats')
        excel_job_stats_ws.protection.sheet = xls_locked

        excel_instance_family_summary_ws = excel_wb.create_sheet(title='InstanceFamilySummary')
        excel_instance_family_summary_ws.protection.sheet = xls_locked

        excel_config_ws = excel_wb.create_sheet(title='Config')
        excel_config_ws.protection.sheet = xls_locked

        excel_instance_info_ws = excel_wb.create_sheet(title='InstanceFamilyRates')
        excel_instance_info_ws.protection.sheet = xls_locked

        excel_hourly_ws = excel_wb.create_sheet(title='Hourly')
        excel_hourly_ws.protection.sheet = xls_locked
        excel_core_hours_chart_ws = excel_wb.create_sheet(title='Core Hours Chart')

        # CostSummary Worksheet
        excel_summary_ws.column_dimensions['A'].width = 35
        excel_summary_ws.column_dimensions['B'].width = 25
        excel_summary_ws.column_dimensions['B'].alignment = XlsAlignment(horizontal='right')
        row = 1
        excel_summary_ws[f'A{row}'] = 'First hour to analyze'
        first_hour_cell = excel_summary_ws[f'B{row}']
        first_hour_cell.value = 0
        first_hour_cell.protection = xls_unlocked
        first_hour_cell_ref = f'CostSummary!$B${row}'
        row += 1
        excel_summary_ws[f'A{row}'] = 'Last hour to analyze'
        last_hour_cell = excel_summary_ws[f'B{row}']
        last_hour_cell.protection = xls_unlocked
        last_hour_cell_ref = f'CostSummary!$B${row}'
        row += 2
        excel_summary_ws.cell(row=row, column=1).value = f'EC2 Savings Plan (ESP) Hourly Commits:'
        esp_hourly_commit_cell_refs = {}
        esp_hourly_commit_first_row = row + 1
        for instance_family in instance_families:
            row += 1
            excel_summary_ws[f'A{row}'] = f'{instance_family}'
            cell = excel_summary_ws[f'B{row}']
            cell.value = 0
            cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
            cell.protection = xls_unlocked
            esp_hourly_commit_cell_refs[instance_family] = f'CostSummary!$B${row}'
        esp_hourly_commit_last_row = row
        row += 1
        excel_summary_ws[f'A{row}'] = 'Total'
        cell = excel_summary_ws[f'B{row}']
        if instance_families:
            cell.value = f"=sum(B{esp_hourly_commit_first_row}:B{esp_hourly_commit_last_row})"
        else:
            cell.value = 0
        cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
        esp_hourly_commit_cell_ref = f"CostSummary!${cell.column_letter}${cell.row}"
        row += 2
        excel_summary_ws[f'A{row}'] = 'CSP Hourly Commit'
        cell = excel_summary_ws[f'B{row}']
        cell.value = 0
        cell.protection = xls_unlocked
        cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
        csp_hourly_commit_cell_ref = f'CostSummary!${cell.column_letter}${cell.row}'
        row += 2
        excel_summary_ws[f'A{row}'] = 'Total Spot'
        total_spot_cell = excel_summary_ws[f'B{row}']
        total_spot_cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
        row += 1
        excel_summary_ws[f'A{row}'] = 'Total OD'
        total_od_cell = excel_summary_ws[f'B{row}']
        total_od_cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
        row += 1
        excel_summary_ws[f'A{row}'] = 'Total ESP'
        total_esp_cell = excel_summary_ws[f'B{row}']
        total_esp_cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
        row += 1
        excel_summary_ws[f'A{row}'] = 'Total CSP'
        total_csp_cell = excel_summary_ws[f'B{row}']
        total_csp_cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
        row += 1
        excel_summary_ws[f'A{row}'] = 'Total'
        total_cell = excel_summary_ws[f'B{row}']
        total_cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
        total_cell_ref = f'CostSummary!${total_cell.column_letter}${total_cell.row}'
        row += 2
        excel_summary_ws.cell(row=row, column=1).value = 'Use Excel Solver to optimize savings plans'
        row += 1
        excel_summary_ws.cell(row=row, column=1).value = 'Enable solver'
        row += 1
        excel_summary_ws.cell(row=row, column=1).value = 'File -> Options'
        row += 1
        excel_summary_ws.cell(row=row, column=1).value = 'Select Add-ins on the left'
        row += 1
        excel_summary_ws.cell(row=row, column=1).value = 'Manage: Excel Add-ins, Click Go'
        row += 1
        excel_summary_ws.cell(row=row, column=1).value = 'Check Solver Add-in, select OK'
        row += 2
        excel_summary_ws.cell(row=row, column=1).value = 'Select Data in menu'
        row += 1
        excel_summary_ws.cell(row=row, column=1).value = 'Select Solver in the Analyze section of the ribbon'
        row += 1
        excel_summary_ws.cell(row=row, column=1).value = f'Set "Set Objective" to Total: {total_cell_ref}'
        row += 1
        excel_summary_ws.cell(row=row, column=1).value = f'Set "To:" to Min'
        row += 1
        excel_summary_ws.cell(row=row, column=1).value = f'Set "By Changing Variable Cells:" to the savings plan commits: $B${esp_hourly_commit_first_row}:$B${esp_hourly_commit_last_row},{csp_hourly_commit_cell_ref}'
        row += 1
        excel_summary_ws.cell(row=row, column=1).value = f'Set "Select a Solving Method:" to GRG Nonlinear'
        row += 1
        excel_summary_ws.cell(row=row, column=1).value = f'Select Solve'

        # JobStats Worksheet
        max_column_widths = {}
        row = 1
        column = 1
        excel_job_stats_ws.cell(row=row, column=column).value = 'Memory Size (GB)'
        max_column_widths[column] = len(excel_job_stats_ws.cell(row=row, column=column).value)
        column += 1
        for runtime in self.get_ranges(self.runtime_ranges_minutes):
            excel_job_stats_ws.cell(row=row, column=column).value = f'{runtime} Minutes'
            excel_job_stats_ws.merge_cells(start_row=row, start_column=column, end_row=row, end_column=column+2)
            excel_job_stats_ws.cell(row=row, column=column).alignment = XlsAlignment(horizontal='center')
            column += 3
        row += 1
        column = 2
        for runtime in self.get_ranges(self.runtime_ranges_minutes):
            excel_job_stats_ws.cell(row=row, column=column).value = 'Job count'
            max_column_widths[column] = len(excel_job_stats_ws.cell(row=row, column=column).value)
            column += 1
            excel_job_stats_ws.cell(row=row, column=column).value = 'Total duration'
            max_column_widths[column] = len(excel_job_stats_ws.cell(row=row, column=column).value)
            column += 1
            excel_job_stats_ws.cell(row=row, column=column).value = 'Total wait time'
            max_column_widths[column] = len(excel_job_stats_ws.cell(row=row, column=column).value)
            column += 1
        row += 1
        for ram in self.get_ranges(self.ram_ranges_GB):
            column = 1
            excel_job_stats_ws.cell(row=row, column=column).value = f"{ram} GB"
            max_column_widths[column] = max(max_column_widths[column], len(excel_job_stats_ws.cell(row=row, column=column).value))
            column += 1
            for runtime in self.get_ranges(self.runtime_ranges_minutes):
                summary = self.job_data_collector[ram][runtime]
                excel_job_stats_ws.cell(row=row, column=column).value = summary['number_of_jobs']
                excel_job_stats_ws.cell(row=row, column=column+1).value = summary['total_duration_minutes']
                excel_job_stats_ws.cell(row=row, column=column+2).value = summary['total_wait_minutes']
                column += 3
            row += 1
        for column, max_column_width in max_column_widths.items():
            excel_job_stats_ws.column_dimensions[xl_get_column_letter(column)].width = max_column_width + 1
        row += 1
        column = 1
        # Add a chart to show the distribution by job characteristics
        # job_count_chart = BarChart3D()
        # job_count_chart.title = 'Job Count by Duration and Memory Size'
        # job_count_chart.style = 13
        # job_count_chart.y_axis.title = 'Core Hours'
        # job_count_chart.x_axis.title = 'Relative Hour'
        # job_count_chart.width = 30
        # job_count_chart.height = 15
        # excel_core_hours_chart_ws.add_chart(job_count_chart, excel_job_stats_ws.cell(row=row, column=column).coordinate)

        # Config Worksheet
        excel_config_ws.column_dimensions['A'].width = 35
        excel_config_ws.column_dimensions['B'].width = 25
        excel_config_ws.column_dimensions['B'].alignment = XlsAlignment(horizontal='right')
        row = 1
        excel_config_ws[f'A{row}'] = 'Region'
        excel_config_ws[f'B{row}'] = self.region
        row += 2
        excel_config_ws[f'A{row}'] = 'Minimum CPU Speed (GHz)'
        excel_config_ws[f'B{row}'] = self.config['consumption_model_mapping']['minimum_cpu_speed']
        row += 1
        excel_config_ws[f'A{row}'] = 'Maximum minutes for spot'
        excel_config_ws[f'B{row}'] = self.config['consumption_model_mapping']['maximum_minutes_for_spot']
        row += 2
        esp_term = f"EC2 SP {self.config['consumption_model_mapping']['ec2_savings_plan_duration']}yr {self.config['consumption_model_mapping']['ec2_savings_plan_payment_option']}"
        excel_config_ws[f'A{row}'] = 'EC2 Savings Plan (ESP) Term'
        excel_config_ws[f'B{row}'] = esp_term
        row += 2
        csp_term = f"Compute SP {self.config['consumption_model_mapping']['ec2_savings_plan_duration']}yr {self.config['consumption_model_mapping']['ec2_savings_plan_payment_option']}"
        excel_config_ws[f'A{row}'] = 'Compute Savings Plan (CSP) Term'
        excel_config_ws[f'B{row}'] = csp_term

        # InstanceFamilyRates Worksheet
        instance_info_headings = ['Instance Family', 'OD Rate', 'ESP Rate','ESP Discount', 'ESP Core*Hr Commit', 'CSP Rate', 'CSP Discount', 'CSP Max Core*Hr Commit']
        instance_family_cols = {}
        instance_family_col_letters = {}
        for instance_info_heading_column, instance_info_heading in enumerate(instance_info_headings, start=1):
            instance_family_cols[instance_info_heading] = instance_info_heading_column
            instance_family_col_letters[instance_info_heading] = xl_get_column_letter(instance_info_heading_column)
            excel_instance_info_ws.cell(row=1, column=instance_info_heading_column, value=instance_info_heading)
            excel_instance_info_ws.column_dimensions[xl_get_column_letter(instance_info_heading_column)].width = len(instance_info_heading) + 1
        instance_family_rows = {}
        csp_discounts = {}
        for instance_family_row, instance_family in enumerate(instance_families, start=2):
            instance_family_rows[instance_family] = instance_family_row
            excel_instance_info_ws.cell(row=instance_family_row, column=1, value=instance_family)
            instance_type = self.instance_family_info[instance_family]['MaxInstanceType']
            coreCount = self.instance_type_info[instance_type]['CoreCount']
            od_rate = self.instance_type_info[instance_type]['pricing']['OnDemand']/coreCount
            excel_instance_info_ws.cell(row=instance_family_row, column=instance_family_cols['OD Rate'], value=od_rate)
            excel_instance_info_ws.cell(row=instance_family_row, column=instance_family_cols['ESP Rate'], value=self.instance_type_info[instance_type]['pricing']['EC2SavingsPlan'][esp_term]/coreCount)
            excel_instance_info_ws.cell(row=instance_family_row, column=instance_family_cols['ESP Discount'], value=f"=({instance_family_col_letters['OD Rate']}{instance_family_row}-{instance_family_col_letters['ESP Rate']}{instance_family_row})/{instance_family_col_letters['OD Rate']}{instance_family_row}")
            excel_instance_info_ws.cell(row=instance_family_row, column=instance_family_cols['ESP Core*Hr Commit'], value=f"={esp_hourly_commit_cell_refs[instance_family]}/{instance_family_col_letters['ESP Rate']}{instance_family_row}")
            csp_rate = self.instance_type_info[instance_type]['pricing']['ComputeSavingsPlan'][csp_term]/coreCount
            excel_instance_info_ws.cell(row=instance_family_row, column=instance_family_cols['CSP Rate'], value=csp_rate)
            csp_discounts[instance_family] = (od_rate - csp_rate)/od_rate
            excel_instance_info_ws.cell(row=instance_family_row, column=instance_family_cols['CSP Discount'], value=f"=({instance_family_col_letters['OD Rate']}{instance_family_row}-{instance_family_col_letters['CSP Rate']}{instance_family_row})/{instance_family_col_letters['OD Rate']}{instance_family_row}")
            excel_instance_info_ws.cell(row=instance_family_row, column=instance_family_cols['CSP Max Core*Hr Commit'], value=f"={csp_hourly_commit_cell_ref}/{instance_family_col_letters['CSP Rate']}{instance_family_row}")

        # CSPs are applied in descending order by size of the discount
        instance_families_by_descending_csp_discounts = sorted(csp_discounts.items(), key=operator.itemgetter(1), reverse=True)
        logger.info(f"instance_families_by_descending_csp_discounts: {instance_families_by_descending_csp_discounts}")

        # Hourly Worksheet
        excel_hourly_ws.freeze_panes = excel_hourly_ws['B2']
        hourly_columns = {}
        hourly_column_letters = {}
        hourly_field_names = ['Relative Hour','Total Spot Costs']
        column = 0
        for field_name in hourly_field_names:
            column += 1
            hourly_column_letters[field_name] = xl_get_column_letter(column)
        hourly_instance_family_field_names = [
            'CHr',
            'ESP CHr',  # The actual number of ESP core hours used. Doesn't affect cost calculation, but can be used to get the ESP utilization ration.
            'CSP CHr',  # The actual number of CSP core hours used. This is necessary since the CSP spans instance families which have different discounts.
            'CSP Cost', # CSP cost for this instance family. This is used to get the total amount of the CSP used so far.
            'OD CHr',   # The OD core hours used. Excess core hours not paid for savings plans.
            'OD Cost']  # OD cost
        for instance_family in instance_families:
            hourly_columns[instance_family] = {}
            hourly_column_letters[instance_family] = {}
            for instance_family_field_name in hourly_instance_family_field_names:
                column += 1
                field_name = f"{instance_family} {instance_family_field_name}"
                hourly_field_names.append(field_name)
                hourly_columns[instance_family][field_name] = column
                hourly_column_letters[field_name] = xl_get_column_letter(column)
                hourly_column_letters[instance_family][instance_family_field_name] = xl_get_column_letter(column)
        hourly_final_field_names = [
            'CSP Cost',     # Total CSP cost. Can be used to calculate CSP utilization
            'OD Cost',      # On demand cost. Don't include ESP and CSP costs because they are fixed per hour
            'Total CHr',    # Total core hours
            'Total OD CHr', # Total on demand core hours. Excludes core hours provided by ESPs and CSPs.
            'Total Cost',   # Total cost. Spot, ESP, CSP, and OD
        ]
        for field_name in hourly_final_field_names:
            column += 1
            hourly_field_names.append(field_name)
            hourly_column_letters[field_name] = xl_get_column_letter(column)
        excel_hourly_ws_columns = len(hourly_field_names)
        for field_column, field_name in enumerate(hourly_field_names, start=1):
            excel_hourly_ws.cell(row=1, column=field_column, value=field_name)
            excel_hourly_ws.column_dimensions[xl_get_column_letter(field_column)].width = len(field_name) + 1
        logger.debug(f"excel_hourly_ws_columns: {excel_hourly_ws_columns}")

        hour_list = list(self.hourly_stats.keys())
        hour_list.sort()
        number_of_hours = 0
        # convert from absolute hour to relative one (obfuscation)
        if hour_list:
            first_hour = hour_list[0]
            last_hour = hour_list[-1]
            logger.info(f"First hour = {first_hour} = {datetime.fromtimestamp(first_hour * SECONDS_PER_HOUR)}")
            logger.info(f"Last  hour = {last_hour} = {datetime.fromtimestamp(last_hour * SECONDS_PER_HOUR)}")
            logger.info(f"{last_hour - first_hour + 1} total hours")
            prev_relative_hour = 0
            for hour in hour_list:
                relative_hour = hour - first_hour
                while prev_relative_hour < relative_hour - 1:
                    # add zero values for all missing hours
                    # Need to add hourly rate for ESP and CSP which are charged whether used or not
                    row = number_of_hours + 2
                    excel_hourly_ws[f"{hourly_column_letters['Relative Hour']}{row}"] = prev_relative_hour + 1
                    excel_hourly_ws[f"{hourly_column_letters['Total Spot Costs']}{row}"] = 0
                    for instance_family in instance_families:
                        for instance_family_field_name in hourly_instance_family_field_names:
                            excel_hourly_ws[f"{hourly_column_letters[instance_family][instance_family_field_name]}{row}"] = 0
                    excel_hourly_ws[f"{hourly_column_letters['CSP Cost']}{row}"] = 0
                    excel_hourly_ws[f"{hourly_column_letters['OD Cost']}{row}"] = 0
                    excel_hourly_ws[f"{hourly_column_letters['Total Cost']}{row}"] = f"={esp_hourly_commit_cell_ref}+{csp_hourly_commit_cell_ref}"

                    number_of_hours += 1
                    prev_relative_hour += 1

                row = number_of_hours + 2
                logger.debug(f"row: {row}")
                excel_hourly_ws.cell(row=row, column=1, value=relative_hour)
                excel_hourly_ws.cell(row=row, column=2, value=self.hourly_stats[hour]['spot'])
                od_cost_formula = '=0'
                csp_cost_total_formula = '0'
                total_core_hour_formula = '=0'
                total_od_core_hour_formula = '=0'
                for instance_family, instance_family_csp_discount in instance_families_by_descending_csp_discounts:
                    instance_family_row = instance_family_rows[instance_family]
                    # Total core hours
                    core_hours = self.hourly_stats[hour]['on_demand']['core_hours'].get(instance_family, 0)
                    excel_hourly_ws[f"{hourly_column_letters[instance_family]['CHr']}{row}"] = core_hours
                    # ESP core hours actually used
                    excel_hourly_ws[f"{hourly_column_letters[instance_family]['ESP CHr']}{row}"] = f"=min({hourly_column_letters[instance_family]['CHr']}{row}, InstanceFamilyRates!${instance_family_col_letters['ESP Core*Hr Commit']}${instance_family_row})"
                    # CSP core hours used by this instance family
                    # First calculate the remaining instance family core hours by subtracting the ESP core hours.
                    # First calculate the remaining CSP commit available.
                    # Then use the available CSP dollars to calculate the number of CSP core hours available
                    # Then use as many of those CSP core hours as possible.
                    excel_hourly_ws[f"{hourly_column_letters[instance_family]['CSP CHr']}{row}"] = f"=min({hourly_column_letters[instance_family]['CHr']}{row}-{hourly_column_letters[instance_family]['ESP CHr']}{row}, ({csp_hourly_commit_cell_ref}-({csp_cost_total_formula}))/InstanceFamilyRates!${instance_family_col_letters['CSP Rate']}${instance_family_row})"
                    # CSP Cost
                    excel_hourly_ws[f"{hourly_column_letters[instance_family]['CSP Cost']}{row}"] = f"={hourly_column_letters[instance_family]['CSP CHr']}{row}*InstanceFamilyRates!${instance_family_col_letters['CSP Rate']}${instance_family_row}"
                    # OD core hours
                    excel_hourly_ws[f"{hourly_column_letters[instance_family]['OD CHr']}{row}"] = f"={hourly_column_letters[instance_family]['CHr']}{row}-{hourly_column_letters[instance_family]['ESP CHr']}{row}-{hourly_column_letters[instance_family]['CSP CHr']}{row}"
                    excel_hourly_ws[f"{hourly_column_letters[instance_family]['OD Cost']}{row}"] = f"={hourly_column_letters[instance_family]['CSP CHr']}{row}*InstanceFamilyRates!$C${instance_family_row}+{hourly_column_letters[instance_family]['OD CHr']}{row}*InstanceFamilyRates!$B${instance_family_row}"
                    csp_cost_total_formula += f"+{hourly_column_letters[instance_family]['CSP Cost']}{row}"
                    od_cost_formula += f"+{hourly_column_letters[instance_family]['OD Cost']}{row}"
                    total_core_hour_formula += f"+{hourly_column_letters[instance_family]['CHr']}{row}"
                    total_od_core_hour_formula += f"+{hourly_column_letters[instance_family]['OD CHr']}{row}"
                excel_hourly_ws[f"{hourly_column_letters['CSP Cost']}{row}"] = f"={csp_cost_total_formula}"
                excel_hourly_ws[f"{hourly_column_letters['OD Cost']}{row}"] = f"{od_cost_formula}"
                excel_hourly_ws[f"{hourly_column_letters['Total CHr']}{row}"] = f"{total_core_hour_formula}"
                excel_hourly_ws[f"{hourly_column_letters['Total OD CHr']}{row}"] = f"{total_od_core_hour_formula}"
                excel_hourly_ws[f"{hourly_column_letters['Total Cost']}{row}"] = f"={hourly_column_letters['Total Spot Costs']}{row}+{esp_hourly_commit_cell_ref}+{csp_hourly_commit_cell_ref}+{hourly_column_letters['OD Cost']}{row}"

                number_of_hours += 1
                prev_relative_hour = relative_hour

        last_hour_cell.value = number_of_hours - 1

        # CostSummary Worksheet
        total_spot_cell.value = f'=sum(indirect("Hourly!{hourly_column_letters["Total Spot Costs"]}" & {first_hour_cell_ref}+2 & ":{hourly_column_letters["Total Spot Costs"]}" & {last_hour_cell_ref}+2))'
        total_esp_cell.value = f'=({last_hour_cell_ref}-{first_hour_cell_ref}+1)*{esp_hourly_commit_cell_ref}'
        total_csp_cell.value = f'=({last_hour_cell_ref}-{first_hour_cell_ref}+1)*{csp_hourly_commit_cell_ref}'
        total_od_cell.value = f'=sum(indirect("Hourly!{hourly_column_letters["OD Cost"]}" & {first_hour_cell_ref}+2 & ":{hourly_column_letters["OD Cost"]}" & {last_hour_cell_ref}+2))'
        total_cell.value =    f'=sum(indirect("Hourly!{hourly_column_letters["Total Cost"]}" & {first_hour_cell_ref}+2 & ":{hourly_column_letters["Total Cost"]}" & {last_hour_cell_ref}+2))'

        # InstanceFamilySummary Worksheet
        row = 1
        column = 1
        cell = excel_instance_family_summary_ws.cell(row=row, column=column)
        cell.value = 'Instance Family'
        excel_instance_family_summary_ws.column_dimensions[cell.column_letter].width = len(cell.value) + 1
        column += 1
        cell = excel_instance_family_summary_ws.cell(row=row, column=column)
        cell.value = 'Min Core Hours'
        excel_instance_family_summary_ws.column_dimensions[cell.column_letter].width = len(cell.value) + 1
        excel_instance_family_summary_ws.column_dimensions[cell.column_letter].alignment = XlsAlignment(horizontal='right')
        column += 1
        cell = excel_instance_family_summary_ws.cell(row=row, column=column)
        cell.value = 'Avg Core Hours'
        excel_instance_family_summary_ws.column_dimensions[cell.column_letter].width = len(cell.value) + 1
        excel_instance_family_summary_ws.column_dimensions[cell.column_letter].alignment = XlsAlignment(horizontal='right')
        column += 1
        cell = excel_instance_family_summary_ws.cell(row=row, column=column)
        cell.value = 'Max Core Hours'
        excel_instance_family_summary_ws.column_dimensions[cell.column_letter].width = len(cell.value) + 1
        excel_instance_family_summary_ws.column_dimensions[cell.column_letter].alignment = XlsAlignment(horizontal='right')
        column += 1
        cell = excel_instance_family_summary_ws.cell(row=row, column=column)
        cell.value = 'Min OD Cost'
        excel_instance_family_summary_ws.column_dimensions[cell.column_letter].width = len(cell.value) + 1
        excel_instance_family_summary_ws.column_dimensions[cell.column_letter].alignment = XlsAlignment(horizontal='right')
        column += 1
        cell = excel_instance_family_summary_ws.cell(row=row, column=column)
        cell.value = 'Avg OD Cost'
        excel_instance_family_summary_ws.column_dimensions[cell.column_letter].width = len(cell.value) + 1
        excel_instance_family_summary_ws.column_dimensions[cell.column_letter].alignment = XlsAlignment(horizontal='right')
        column += 1
        cell = excel_instance_family_summary_ws.cell(row=row, column=column)
        cell.value = 'Max OD Cost'
        excel_instance_family_summary_ws.column_dimensions[cell.column_letter].width = len(cell.value) + 1
        excel_instance_family_summary_ws.column_dimensions[cell.column_letter].alignment = XlsAlignment(horizontal='right')

        instance_family_first_row = row + 1
        for instance_family in instance_families:
            row += 1
            column = 1
            excel_instance_family_summary_ws.cell(row=row, column=column).value = f'{instance_family}'

            column += 1
            cell = excel_instance_family_summary_ws.cell(row=row, column=column)
            cell.value = f'=min(indirect("Hourly!{hourly_column_letters[instance_family + " CHr"]}" & CostSummary!{first_hour_cell.coordinate}+2 & ":{hourly_column_letters[instance_family + " CHr"]}" & CostSummary!{last_hour_cell.coordinate}+2))'
            column += 1
            cell = excel_instance_family_summary_ws.cell(row=row, column=column)
            cell.value = f'=average(indirect("Hourly!{hourly_column_letters[instance_family + " CHr"]}" & CostSummary!{first_hour_cell.coordinate}+2 & ":{hourly_column_letters[instance_family + " CHr"]}" & CostSummary!{last_hour_cell.coordinate}+2))'
            column += 1
            cell = excel_instance_family_summary_ws.cell(row=row, column=column)
            cell.value = f'=max(indirect("Hourly!{hourly_column_letters[instance_family + " CHr"]}" & CostSummary!{first_hour_cell.coordinate}+2 & ":{hourly_column_letters[instance_family + " CHr"]}" & CostSummary!{last_hour_cell.coordinate}+2))'

            column += 1
            cell = excel_instance_family_summary_ws.cell(row=row, column=column)
            cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
            cell.value = f'=min(indirect("Hourly!{hourly_column_letters[instance_family + " OD Cost"]}" & CostSummary!{first_hour_cell.coordinate}+2 & ":{hourly_column_letters[instance_family + " OD Cost"]}" & CostSummary!{last_hour_cell.coordinate}+2))'
            column += 1
            cell = excel_instance_family_summary_ws.cell(row=row, column=column)
            cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
            cell.value = f'=average(indirect("Hourly!{hourly_column_letters[instance_family + " OD Cost"]}" & CostSummary!{first_hour_cell.coordinate}+2 & ":{hourly_column_letters[instance_family + " OD Cost"]}" & CostSummary!{last_hour_cell.coordinate}+2))'
            column += 1
            cell = excel_instance_family_summary_ws.cell(row=row, column=column)
            cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
            cell.value = f'=max(indirect("Hourly!{hourly_column_letters[instance_family + " OD Cost"]}" & CostSummary!{first_hour_cell.coordinate}+2 & ":{hourly_column_letters[instance_family + " OD Cost"]}" & CostSummary!{last_hour_cell.coordinate}+2))'
        instance_family_last_row = row
        row += 1
        excel_instance_family_summary_ws.cell(row=row, column=1).value = 'Total'
        for column in range(2, 8):
            cell = excel_instance_family_summary_ws.cell(row=row, column=column)
            if instance_families:
                cell.value = f"=sum({xl_get_column_letter(column)}{instance_family_first_row}:{xl_get_column_letter(column)}{instance_family_last_row})"
            else:
                cell.value = 0
            if column > 4:
                cell.number_format = FORMAT_CURRENCY_USD_SIMPLE

        # Core Hours Charts

        column = 1
        row = 1

        # Stacked Core Hours Chart
        # Stacked line chart with the number of core hours per instance family
        core_hours_chart = XlLineChart()
        core_hours_chart.title = 'Core Hours by Instance Family'
        core_hours_chart.style = 13
        core_hours_chart.y_axis.title = 'Core Hours'
        core_hours_chart.x_axis.title = 'Relative Hour'
        core_hours_chart.grouping = 'stacked'
        core_hours_chart.width = 30
        core_hours_chart.height = 15
        for instance_family in instance_families:
            column = hourly_columns[instance_family][f'{instance_family} CHr']
            data_series = XlReference(excel_hourly_ws, min_col=column, min_row=1, max_col=column, max_row=last_hour_cell.value + 2)
            core_hours_chart.add_data(data_series, titles_from_data=True)
        excel_core_hours_chart_ws.add_chart(core_hours_chart, 'A1')
        row += 30

        # Core Hours Chart by instance family
        for instance_family in instance_families:
            core_hours_chart = XlLineChart()
            core_hours_chart.title = f'{instance_family} Core Hours'
            core_hours_chart.style = 13
            core_hours_chart.y_axis.title = 'Core Hours'
            core_hours_chart.x_axis.title = 'Relative Hour'
            core_hours_chart.width = 30
            core_hours_chart.height = 15
            column = hourly_columns[instance_family][f'{instance_family} CHr']
            data_series = XlReference(excel_hourly_ws, min_col=column, min_row=1, max_col=column, max_row=last_hour_cell.value + 2)
            core_hours_chart.add_data(data_series, titles_from_data=True)
            cell = excel_core_hours_chart_ws.cell(row=row, column=1)
            excel_core_hours_chart_ws.add_chart(core_hours_chart, cell.coordinate)
            row += 30

        excel_wb.save(hourly_stats_xlsx)

    def analyze_jobs(self):
        '''
        Analyze jobs

        Analyze jobs 1 by 1.
        Select a pricing plan and an instance type for each job and calculate the job's cost.
        Bin jobs by the hour in which they started and save the hourly jobs in separate CSV files.
        Process the hourly files to accumulate the hourly on-demand and spot costs broken out by instance family.
        Write the hourly costs into a separate CSV file.
        Write an overall job summary.

        Scalability is a key consideration because millions of jobs must be processessed so the analysis cannot be
        done in memory.
        First breaking the jobs out into hourly chunks makes the process scalable.
        '''
        if not self.instance_type_info:
            self.get_instance_type_info()

        self._clear_job_stats()   # allows calling multiple times in testing
        self._cleanup_hourly_files()

        total_jobs = 0
        total_failed_jobs = 0
        while True:
            job = self._scheduler_parser.parse_job()
            if not job:
                logger.debug(f"No more jobs")
                break
            if not self._filter_job_queue(job):
                logger.debug(f"Job {job.job_id} filtered out queue {job.queue}")
                continue
            if not self._filter_job_project(job):
                logger.debug(f"Job {job.job_id} filtered out project {job.project}")
                continue
            total_jobs += 1
            if self._scheduler_parser._output_csv:
                self._scheduler_parser.write_job_to_csv(job)
            try:
                job_cost_data = self.analyze_job(job)
            except RuntimeError as e:
                total_failed_jobs += 1
                logger.error(f"{e}")
                continue
            self._add_job_to_collector(job)
            self._add_job_to_hourly_bucket(job_cost_data)
        logger.info(f"Finished processing {total_jobs-total_failed_jobs}/{total_jobs} jobs")

        # Dump pending jobs and summary to output files
        self._write_hourly_jobs_buckets_to_file()
        self._dump_job_collector_to_csv()

        self._process_hourly_jobs()
        self._write_hourly_stats()

    def _filter_job_queue(self, job: SchedulerJobInfo) -> bool:
        '''
        Filter the job queue

        Args:
            job (SchedulerJobInfo): Parsed job information
        Returns:
            bool: True if the job should be analyzed
        '''
        logger.debug(f"Filtering job {job.job_id} queue {job.queue}")
        if job.queue == None:
            queue = ''
        else:
            queue = job.queue
        for (include_filter, filter_regexp) in self._queue_filter_regexps:
            if filter_regexp.match(queue):
                logger.debug(f"job {job.job_id} queue={queue} matched {filter_regexp} include={include_filter}")
                return include_filter
        logger.debug(f"job {job.job_id} queue={queue} didn't match any filters")
        return False

    def _filter_job_project(self, job: SchedulerJobInfo) -> bool:
        '''
        Filter the job project

        Args:
            job (SchedulerJobInfo): Parsed job information
        Returns:
            bool: True if the job should be analyzed
        '''
        logger.debug(f"Filtering job {job.job_id} project {job.project}")
        if job.project == None:
            project = ''
        else:
            project = job.project
        for (include_filter, filter_regexp) in self._project_filter_regexps:
            if filter_regexp.match(project):
                logger.debug(f"job {job.job_id} project={project} matched {filter_regexp} include={include_filter}")
                return include_filter
        logger.debug(f"job {job.job_id} project={project} didn't match any filters.")
        return False

    def analyze_job(self, job: SchedulerJobInfo) -> JobCost:
        '''
        process a single job

        Args:
            job (SchedulerJobInfo): Parsed job information
        Returns:
            JobCost: Job cost information
        '''
        # Find the right instance type to run the job + its price
        num_hosts = job.num_hosts
        min_memory_per_instance = ceil(job.max_mem_gb / num_hosts)
        num_cores_per_instance = ceil(job.num_cores / num_hosts)
        potential_instance_types = self.get_instance_by_spec(min_memory_per_instance, num_cores_per_instance, self.minimum_cpu_speed)
        if len(potential_instance_types) == 0:
            raise RuntimeError(f"Job {job.job_id} with {min_memory_per_instance} GB and {num_cores_per_instance} cores is too big to fit in a single instance.")
        job_runtime_minutes = job.run_time_td.total_seconds()/60
        spot_threshold = self.config['consumption_model_mapping']['maximum_minutes_for_spot']
        spot = job_runtime_minutes < spot_threshold
        (instance_type, rate) = self.get_lowest_priced_instance(potential_instance_types, spot)
        if spot:
            (on_demand_instance_type, on_demand_rate) = self.get_lowest_priced_instance(potential_instance_types, spot=False)
            if not instance_type or (on_demand_rate < rate):
                # No spot pricing available. Get lowest cost on-demand instance type
                spot = False
                instance_type = on_demand_instance_type
                rate = on_demand_rate
        logger.debug(f"Lowest priced instance type: {instance_type} spot={spot} rate={rate}")
        instance_family = EC2InstanceTypeInfo.get_instance_family(instance_type)
        job_cost_data = JobCost(job, spot, instance_family, instance_type, rate)
        return job_cost_data

def main():
    try:
        parser = argparse.ArgumentParser(description="Analyze jobs", formatter_class=argparse.ArgumentDefaultsHelpFormatter)
        parser.add_argument("--starttime", help="Select jobs after the specified time. Format YYYY-MM-DDTHH:MM:SS")
        parser.add_argument("--endtime", help="Select jobs before the specified time. Format YYYY-MM-DDTHH:MM:SS")
        parser.add_argument("--config", required=False, default=f'{dirname(__file__)}/config.yml', help="Configuration file.")
        parser.add_argument("--acknowledge-config", required=False, action='store_const', const=True, default=False, help="Acknowledge configuration file contents so don't get prompt.")
        parser.add_argument("--output-dir", required=False, default="output", help="Directory where output will be written")
        parser.add_argument("--output-csv", required=False, default=None, help="CSV file with parsed job completion records")

        # Subparsers for scheduler specific arguments
        # required=True requires python 3.7 or later. Remove to make the installation simpler.
        subparsers = parser.add_subparsers(metavar='parser', dest='parser', help=f'Choose the kind of information to parse. {__file__} <parser> -h for parser specific arguments.')

        accelerator_parser = subparsers.add_parser('accelerator', help='Parse Accelerator (nc) job information', formatter_class=argparse.ArgumentDefaultsHelpFormatter)
        accelerator_parser.add_argument("--default-mem-gb", required=False, default=AcceleratorLogParser.DEFAULT_MEMORY_GB, help="Default amount of memory (in GB) requested for jobs.")
        accelerator_mutex_group = accelerator_parser.add_mutually_exclusive_group(required=True)
        accelerator_mutex_group.add_argument("--sql-output-file", help=f"File where the output of sql query will be written. Cannot be used with --sql-input-file. Required if --sql-input-file not set. \nCommand to create file:\n{AcceleratorLogParser._VOVSQL_QUERY_COMMAND} > SQL_OUTPUT_FILE")
        accelerator_mutex_group.add_argument("--sql-input-file", help="File with the output of sql query so can process it offline. Cannot be used with --sql-output-file. Required if --sql-output-file not set.")

        csv_parser = subparsers.add_parser('csv', help='Parse CSV from already parsed job information.', formatter_class=argparse.ArgumentDefaultsHelpFormatter)
        csv_parser.add_argument("--input-csv", required=True, help="CSV file with parsed job info from scheduler parser.")

        lsf_parser = subparsers.add_parser('lsf', help='Parse LSF logfiles', formatter_class=argparse.ArgumentDefaultsHelpFormatter)
        lsf_parser.add_argument("--logfile-dir", required=False, help="LSF logfile directory")
        lsf_parser.add_argument("--default-max-mem-gb", type=float, required=True, help="Default maximum memory for a job in GB.")

        slurm_parser = subparsers.add_parser('slurm', help='Parse Slurm job information', formatter_class=argparse.ArgumentDefaultsHelpFormatter)
        slurm_parser.add_argument("--slurm-root", required=False, help="Directory that contains the Slurm bin directory.")
        slurm_mutex_group = slurm_parser.add_mutually_exclusive_group()
        slurm_mutex_group.add_argument("--sacct-output-file", required=False, help="File where the output of sacct will be written. Cannot be used with --sacct-input-file. Required if --sacct-input-file not set.")
        slurm_mutex_group.add_argument("--sacct-input-file", required=False, help="File with the output of sacct so can process sacct output offline. Cannot be used with --sacct-output-file. Required if --sacct-output-file not set.")

        hourly_stats_parser = subparsers.add_parser('hourly_stats', help='Parse hourly_*.csv hourly files so can create Excel workbook (xlsx).', formatter_class=argparse.ArgumentDefaultsHelpFormatter)

        hourly_stats_csv_parser = subparsers.add_parser('hourly_stats_csv', help='Parse hourly_stats.csv file so can create Excel workbook (xlsx).', formatter_class=argparse.ArgumentDefaultsHelpFormatter)
        hourly_stats_csv_parser.add_argument("--input-hourly-stats-csv", required=True, help="Existing hourly_stats.csv file to use as input.")

        parser.add_argument("--queues", default=None, help="Comma separated list of regular expressions of queues to include/exclude. Prefix the queue with '-' to exclude. The regular expressions are evaluated in the order given and the first match has precedence and stops further evaluations. Regular expressions have an implicit ^ at the beginning.")

        parser.add_argument("--projects", default=None, help="Comma separated list of regular expressions of projects to include/exclude. Prefix the project with '-' to exclude. The regular expressions are evaluated in the order given and the first match has precedence and stops further evaluations. Regular expressions have an implicit ^ at the beginning.")

        parser.add_argument("--disable-version-check", action='store_const', const=True, default=False, help="Disable git version check")

        parser.add_argument("--debug", '-d', action='store_const', const=True, default=False, help="Enable debug mode")
        args = parser.parse_args()

        if args.debug:
            logger.setLevel(logging.DEBUG)
            AcceleratorLogParser_logger.setLevel(logging.DEBUG)
            CSVLogParser_logger.setLevel(logging.DEBUG)
            LSFLogParser_logger.setLevel(logging.DEBUG)
            SchedulerJobInfo_logger.setLevel(logging.DEBUG)
            SchedulerLogParser_logger.setLevel(logging.DEBUG)
            SlurmLogParser_logger.setLevel(logging.DEBUG)
            VersionCheck_logger.setLevel(logging.DEBUG)

        if not args.disable_version_check and not VersionCheck().check_git_version():
            exit(1)

        if not args.parser:
            logger.error("The following arguments are required: parser")
            exit(2)
        logger.info('Started job analyzer')

        if args.parser == 'csv':
            logger.info(f"Reading job data from {args.input_csv}")
            scheduler_parser = CSVLogParser(args.input_csv, args.output_csv, args.starttime, args.endtime)
        elif args.parser == 'accelerator':
            logger.info(f"Parsing Altair Accelerator jobs")
            if args.sql_input_file:
                logger.info(f"    Parsing job results from {args.sql_input_file}")
            else:
                logger.info(f"    Querying job information from Accelerator sql database.")
            scheduler_parser = AcceleratorLogParser(default_mem_gb=float(args.default_mem_gb), sql_input_file=args.sql_input_file, sql_output_file=args.sql_output_file, output_csv=args.output_csv, starttime=args.starttime, endtime=args.endtime)
        elif args.parser == 'lsf':
            if not args.logfile_dir or not args.output_csv:
                logger.error(f"You must provide --logfile-dir and --output-csv for LSF.")
                exit(1)
            logger.info(f"Parsing LSF jobs")
            scheduler_parser = LSFLogParser(args.logfile_dir, args.output_csv, args.default_max_mem_gb, args.starttime, args.endtime)
        elif args.parser == 'slurm':
            logger.info(f"Parsing Slurm jobs")
            if args.sacct_output_file:
                logger.info(f"    Parsing job results from {args.sacct_output_file}")
            else:
                logger.info(f"    Querying job information from Slurm results database")
            scheduler_parser = SlurmLogParser(args.sacct_input_file, args.sacct_output_file, args.output_csv, args.starttime, args.endtime)
        elif args.parser == 'hourly_stats':
            logger.info(f"Parsing hourly jobs from {args.output_dir}/hourly-*.csv")
            scheduler_parser = None
            jobAnalyzer = JobAnalyzer(scheduler_parser, args.config, args.output_dir, args.starttime, args.endtime, queue_filters=args.queues, project_filters=args.projects)
            jobAnalyzer._process_hourly_jobs()
            jobAnalyzer._write_hourly_stats()
        elif args.parser == 'hourly_stats_csv':
            logger.info(f"Parsing {args.output_dir}/hourly_stats.csv")
            scheduler_parser = None
            jobAnalyzer = JobAnalyzer(scheduler_parser, args.config, args.output_dir, args.starttime, args.endtime, queue_filters=args.queues, project_filters=args.projects)
            jobAnalyzer.parse_hourly_stats_csv(args.input_hourly_stats_csv)
            jobAnalyzer._write_hourly_stats()

        if scheduler_parser:
            if args.output_csv:
                logger.info(f"Writing job data to {args.output_csv}")

            jobAnalyzer = JobAnalyzer(scheduler_parser, args.config, args.output_dir, args.starttime, args.endtime, queue_filters=args.queues, project_filters=args.projects)

            # Print out configuration information
            logger.info(f"""Configuration:
            {json.dumps(jobAnalyzer.config, indent=4)}""")
            acknowledge_config = args.acknowledge_config
            while not acknowledge_config:
                print(f"\nIs the correct configuration? (y/n) ")
                answer = input().lower()
                if answer == 'n':
                    exit(1)
                elif answer == 'y':
                    acknowledge_config = True

            jobAnalyzer.analyze_jobs()
    except Exception as e:
        logger.exception(f"Unhandled exception")
        exit(1)

if __name__ == '__main__':
    main()
